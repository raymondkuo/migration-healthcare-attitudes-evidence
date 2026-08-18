# -*- coding: utf-8 -*-
"""Carry the NOT USED reclassification into the published workbook's Known_issues
sheet, so the sheet and data/known_issues.csv cannot drift apart. Patches the two
changed cells in place rather than rebuilding the workbook, which would undo the
traceability work that ran after the builder."""
import os
import openpyxl
import pandas as pd

SITE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(SITE, 'data', 'FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx')

k = pd.read_csv(os.path.join(SITE, 'data', 'known_issues.csv'))
src = k[k.scope.astype(str).str.contains('immigration_country_year', na=False)].iloc[0]


def ascii_ise(s):
    """The sheet is written in plain ASCII; keep it that way."""
    return (str(s).replace('—', '-').replace('–', '-')
            .replace('’', "'").replace('‘', "'")
            .replace('“', '"').replace('”', '"'))


wb = openpyxl.load_workbook(XLSX)
ws = wb['Known_issues']
hdr = [c.value for c in ws[1]]
c_sev, c_scope, c_act = hdr.index('severity') + 1, hdr.index('scope') + 1, hdr.index('action') + 1

hits = [r for r in range(2, ws.max_row + 1)
        if 'immigration_country_year' in str(ws.cell(r, c_scope).value or '')]
assert len(hits) == 1, 'expected one workbook-scoped row, found %d' % len(hits)
row = hits[0]
assert ws.cell(row, c_sev).value == 'HIGH', 'unexpected severity %r' % ws.cell(row, c_sev).value

ws.cell(row, c_sev).value = src['severity']
ws.cell(row, c_act).value = ascii_ise(src['action'])
wb.save(XLSX)

# read back and confirm against the CSV
chk = openpyxl.load_workbook(XLSX, read_only=True)['Known_issues']
got = list(chk.iter_rows(min_row=row, max_row=row, values_only=True))[0]
print('row %d patched' % row)
print('  severity:', got[c_sev - 1])
print('  action  :', str(got[c_act - 1])[:120] + '...')
assert got[c_sev - 1] == src['severity']
assert got[c_act - 1] == ascii_ise(src['action'])
print('\nworkbook sheet matches data/known_issues.csv')
