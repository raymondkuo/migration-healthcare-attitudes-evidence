# -*- coding: utf-8 -*-
"""Carry the source rework into known_issues and the published workbook.

Three issue entries described the old presentation, and one of them was wrong on a
detail: the Italy entry claimed ISMU's series confirmed *every* Italian value, but that
series ends at 2021 — 2022 rests on the XXVIII Rapporto. Corrected here."""
import os
import openpyxl
import pandas as pd

SITE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(SITE, 'data')
XLSX = os.path.join(D, 'FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx')

ACTIONS = {
 ('Switzerland', 'irregular_stock'):
    'The register now cites the source the value is verified against: the SRF report on the '
    'SEM study\'s release of 25 April 2016, which states 76,000. The moved SEM PDF is kept '
    'in the superseded_source_url column. The study\'s own range was 58,000-105,000.',
 ('Italy', 'sources'):
    'The register now cites ISMU\'s own machine-readable series, which reproduces every '
    'Italian value for 2010-2021 exactly. That series ends at 2021, so 2022 (506,000) is '
    'cited to the XXVIII Rapporto ISMU 2022, which states it against 519,000 for the year '
    'before. The blocked press releases are kept in the superseded_source_url column.',
 ('Japan', 'irregular_proxy_overstayers'):
    'No impact on any published number. The 2014 figure of 59,061 is printed in the '
    'Immigration Services Agency document archived here (checked in the archived copy, not '
    'inferred), so the dead mirror is redundant. It is kept in the superseded_source_url '
    'column rather than discarded.',
}

k = pd.read_csv(os.path.join(D, 'known_issues.csv'))
for (scope, var), action in ACTIONS.items():
    m = (k.scope == scope) & (k.variable == var)
    assert m.sum() == 1, '%s / %s -> %d rows' % (scope, var, m.sum())
    k.loc[m, 'action'] = action
    print('%-12s %-28s %s...' % (scope, var, action[:66]))
k.to_csv(os.path.join(D, 'known_issues.csv'), index=False, encoding='utf-8-sig')


def ascii_ise(s):
    return (str(s).replace('—', '-').replace('–', '-').replace('’', "'")
            .replace('‘', "'").replace('“', '"').replace('”', '"'))


SHEETS = {'Known_issues': 'known_issues.csv',
          'Source_register': 'source_register.csv',
          'Panel_final': 'panel_final.csv'}

wb = openpyxl.load_workbook(XLSX)
print()
for sheet, csv in SHEETS.items():
    df = pd.read_csv(os.path.join(D, csv)).fillna('')
    pos = wb.sheetnames.index(sheet)
    before = (wb[sheet].max_row - 1, wb[sheet].max_column)
    del wb[sheet]
    ws = wb.create_sheet(sheet, pos)
    ws.append([ascii_ise(c) for c in df.columns])
    for rec in df.itertuples(index=False, name=None):
        ws.append([ascii_ise(x) if isinstance(x, str) else x for x in rec])
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col[:200] if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 62)
    print('%-18s %d x %d  ->  %d x %d'
          % (sheet, before[0], before[1], ws.max_row - 1, ws.max_column))

# README counts that the rework moved
reg = pd.read_csv(os.path.join(D, 'source_register.csv')).fillna('')
n_sup = int((reg.superseded_source_url.astype(str).str.startswith('http')).sum())
README = {
    'Document source citations':
        '%d source rows across %d distinct URLs' % (len(reg), reg.source_url.nunique()),
    '  archived in the country folders':
        '%d of %d (%d verified live via API, %d archived as documents)'
        % (int((reg.local_file.astype(str).str.len() > 3).sum()), len(reg),
           int((reg.retrieval == 'VERIFIED_API').sum()),
           int((reg.retrieval == 'ARCHIVED').sum())),
    '  not retrievable by any means':
        '0 - every citation relied on is archived here. %d original citations had gone dead '
        'and were replaced by a live source carrying the same figure; each original is kept '
        'in superseded_source_url.' % n_sup,
}
ws = wb['README']
seen = set()
for row in range(1, ws.max_row + 1):
    lab = str(ws.cell(row, 1).value or '')
    if lab in README:
        ws.cell(row, 2).value = ascii_ise(README[lab])
        seen.add(lab)
assert set(README) == seen, 'README labels missing: %s' % (set(README) - seen)
wb.save(XLSX)

print()
chk = openpyxl.load_workbook(XLSX, read_only=True)
for sheet, csv in SHEETS.items():
    df = pd.read_csv(os.path.join(D, csv))
    ws = chk[sheet]
    assert ws.max_row - 1 == len(df) and ws.max_column == len(df.columns), sheet
    print('  %-18s %d x %d  matches %s' % (sheet, ws.max_row - 1, ws.max_column, csv))
print('\nsuperseded citations preserved: %d' % n_sup)
