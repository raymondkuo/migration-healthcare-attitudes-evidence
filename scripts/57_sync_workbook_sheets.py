# -*- coding: utf-8 -*-
"""Rewrite every workbook sheet that has a CSV counterpart from that CSV.

Sheets had drifted from the data files twice (Verification_log and Source_register
predated the Eurostat re-test; Known_issues and Codebook then gained translation
columns). Run this after any change to data/*.csv rather than patching sheets by hand.
README counters that quote row counts are refreshed too."""
import os
import openpyxl
import pandas as pd

SITE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(SITE, 'data')
XLSX = os.path.join(D, 'FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx')

SHEETS = {
    'Panel_final': 'panel_final.csv',
    'Data_quality': 'data_quality.csv',
    'Corrections_applied': 'corrections_applied.csv',
    'Known_issues': 'known_issues.csv',
    'Verification_log': 'verification_log.csv',
    'Source_register': 'source_register.csv',
    'Irregular_estimates_all': 'irregular_estimates_all.csv',
    'Codebook': 'codebook.csv',
    'Deleted_values': 'deleted_values.csv',
}

# The sheet is written in plain ASCII, but the translation columns are Chinese by
# definition — only fold the typographic punctuation that has an ASCII equivalent.
PUNCT = {'—': '-', '–': '-', '’': "'", '‘': "'", '“': '"', '”': '"'}


def clean(x):
    if not isinstance(x, str):
        return x
    for a, b in PUNCT.items():
        x = x.replace(a, b)
    return x


wb = openpyxl.load_workbook(XLSX)
missing = [s for s in SHEETS if s not in wb.sheetnames]
assert not missing, 'workbook has no sheet(s): %s' % missing

for sheet, csv in SHEETS.items():
    fp = os.path.join(D, csv)
    if not os.path.exists(fp):
        print('%-24s SKIP (no %s)' % (sheet, csv))
        continue
    df = pd.read_csv(fp).fillna('')
    pos = wb.sheetnames.index(sheet)
    before = (wb[sheet].max_row - 1, wb[sheet].max_column)
    del wb[sheet]
    ws = wb.create_sheet(sheet, pos)
    ws.append([clean(c) for c in df.columns])
    for rec in df.itertuples(index=False, name=None):
        ws.append([clean(x) for x in rec])
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col[:200] if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 62)
    flag = '' if before == (ws.max_row - 1, ws.max_column) else '   <- changed'
    print('%-24s %d x %-3d -> %d x %d%s'
          % (sheet, before[0], before[1], ws.max_row - 1, ws.max_column, flag))

log = pd.read_csv(os.path.join(D, 'verification_log.csv'))
ws = wb['README']
for row in range(1, ws.max_row + 1):
    if str(ws.cell(row, 1).value or '') == 'Verification_log':
        ws.cell(row, 2).value = ('All %d value-by-value comparisons against live sources, each '
                                 'stamped with the stage and the date it was performed.' % len(log))
wb.save(XLSX)

chk = openpyxl.load_workbook(XLSX, read_only=True)
print()
bad = 0
for sheet, csv in SHEETS.items():
    fp = os.path.join(D, csv)
    if not os.path.exists(fp):
        continue
    df = pd.read_csv(fp)
    ws = chk[sheet]
    ok = (ws.max_row - 1 == len(df) and ws.max_column == len(df.columns))
    bad += not ok
    print('  %-24s %-12s csv %-12s %s'
          % (sheet, '%d x %d' % (ws.max_row - 1, ws.max_column),
             '%d x %d' % (len(df), len(df.columns)), 'ok' if ok else '*** DRIFT ***'))
assert not bad, '%d sheet(s) still drift' % bad
print('\nall %d sheets match data/*.csv' % len(SHEETS))
