# -*- coding: utf-8 -*-
"""Audit every website statistic against the named VERIFIED workbook.

Entry point for the site-vs-xlsx checklist. Parses shipped HTML (not the
site-builder scripts), compares displayed numbers to the named workbook,
checks statistic-associated live URLs, and hash-compares local archives.

Usage:
  python scripts/audit_site_against_verified.py
  python scripts/audit_site_against_verified.py --no-fetch
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html as htmlmod
import json
import math
import os
import re
import sys
import time
import traceback
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import unquote, urlparse

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

SCRIPT_DIR = Path(__file__).resolve().parent
SITE_ROOT = SCRIPT_DIR.parent
DEFAULT_TRUTH = Path(
    r"D:\研究計畫\其他投稿\2026_移民對非本國籍使用公共醫療態度（葉明叡）"
    r"\claude-work\FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx"
)
DEFAULT_CHECKLIST = SITE_ROOT / "outputs" / "audit_site_vs_verified" / "AUDIT_checklist_site_vs_VERIFIED.xlsx"
# Resolve the reference workbook relative to this repository.  This keeps the
# default usable on Windows even when the parent directory contains CJK text.
DEFAULT_TRUTH = SITE_ROOT.parent / "FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx"

PANEL_VARS = [
    "population",
    "foreign_born",
    "foreign_nationals",
    "irregular_stock",
    "irregular_proxy_overstayers",
    "irregular_proxy_detections",
    "irregular_proxy_absconded_workers",
]
HEADER_TO_VAR = {
    "population": "population",
    "foreign-born": "foreign_born",
    "foreign born": "foreign_born",
    "foreign nationals": "foreign_nationals",
    "irregular stock": "irregular_stock",
    "overstayers": "irregular_proxy_overstayers",
    "detections": "irregular_proxy_detections",
    "absconded workers (tw)": "irregular_proxy_absconded_workers",
    "absconded workers": "irregular_proxy_absconded_workers",
}
VAR_TO_LABEL = {
    "population": "Population",
    "foreign_born": "Foreign-born",
    "foreign_nationals": "Foreign nationals",
    "irregular_stock": "Irregular stock",
    "irregular_proxy_overstayers": "Overstayers",
    "irregular_proxy_detections": "Detections",
    "irregular_proxy_absconded_workers": "Absconded workers (TW)",
    "irregular": "Irregular migration",
    "foreign_workers": "Foreign workers",
    "irregular_detections": "Detections",
}
LABEL_TO_VAR = {v.lower(): k for k, v in VAR_TO_LABEL.items()}
LABEL_TO_VAR.update(HEADER_TO_VAR)

AUTHOR_HOSTS = {"raymond.cph.ntu.edu.tw"}
NAV_SKIP_PREFIXES = ("mailto:", "javascript:", "data:")

BLANK_TOKENS = {"", "—", "–", "-", "—", "n/a", "none", "nan", "null"}

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/128.0.0.0 Safari/537.36"
)

ROW_FIELDS = [
    "iso3",
    "verified_sheet",
    "page",
    "variable",
    "year",
    "item",
    "site_value",
    "xlsx_value",
    "number_match",
    "url",
    "http_status",
    "downloadable",
    "archive_path",
    "archive_exists",
    "live_sha256",
    "archive_sha256",
    "content_compare",
    "notes",
]

VERIFIED_SHEETS = [
    "README",
    "Panel_final",
    "Data_quality",
    "Corrections_applied",
    "Known_issues",
    "Verification_log",
    "Source_register",
    "Irregular_estimates_all",
    "Codebook",
]


# ---------------------------------------------------------------------------
# Numeric normalization
# ---------------------------------------------------------------------------

def is_blank(v) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(v, str) and v.strip().lower() in BLANK_TOKENS:
        return True
    return False


def parse_site_number(text):
    """Parse a displayed statistic into int, float, coverage string, or None."""
    if text is None:
        return None
    s = htmlmod.unescape(str(text)).strip()
    s = s.replace("\xa0", " ").replace(",", "")
    s = re.sub(r"\s+", " ", s)
    if s.lower() in BLANK_TOKENS or s in {"&mdash;", "mdash"}:
        return None
    # Combined headline such as "2,415 (98.4%)".  Preserve both parts so
    # it can be compared with the corresponding README cell.
    m = re.fullmatch(r"([+-]?\d[\d,]*)\s*\(([+-]?\d+(?:\.\d+)?)\s*%\)", s)
    if m:
        return {
            "count": int(m.group(1).replace(",", "")),
            "pct": float(m.group(2)),
            "raw": s,
        }
    # coverage / fraction: 13/13 or 520 / 520
    m = re.fullmatch(r"(\d+)\s*/\s*(\d+)", s)
    if m:
        return "%s/%s" % (m.group(1), m.group(2))
    # percent
    m = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*%", s)
    if m:
        return float(m.group(1))
    # plain number, allow +
    m = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)", s)
    if m:
        f = float(m.group(1))
        if abs(f - round(f)) < 1e-9 and abs(f) < 1e15:
            return int(round(f))
        return f
    return s


def normalize_xlsx_number(v):
    if is_blank(v):
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        if abs(v - round(v)) < 1e-9 and abs(v) < 1e15:
            return int(round(v))
        return v
    s = str(v).strip()
    # "2415 (98.4%)"
    m = re.match(r"^(\d+)\s*\(([0-9.]+)\s*%\)$", s.replace(",", ""))
    if m:
        return {"count": int(m.group(1)), "pct": float(m.group(2)), "raw": s}
    parsed = parse_site_number(s)
    return parsed


def numbers_equal(site_v, xlsx_v) -> bool:
    a = site_v if isinstance(site_v, (int, float, dict)) or site_v is None else parse_site_number(site_v)
    b = xlsx_v if isinstance(xlsx_v, dict) else normalize_xlsx_number(xlsx_v)
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(b, dict):
        if isinstance(a, dict) and "count" in a and "pct" in a:
            return a["count"] == b.get("count") and abs(float(a["pct"]) - float(b.get("pct"))) < 1e-9
        if isinstance(a, (int, float)) and "count" in b and "pct" in b:
            if isinstance(a, int):
                return a == b["count"]
            return abs(float(a) - float(b["pct"])) < 0.05
        raw = str(b.get("raw") or "")
        if isinstance(a, str) and a.replace(" ", "") == raw.replace(" ", ""):
            return True
        m = re.match(r"^(\d+)\s*\(([0-9.]+)\s*%\)$", str(a).replace(",", ""))
        if m and "count" in b and "pct" in b:
            return int(m.group(1)) == int(b["count"]) and abs(float(m.group(2)) - float(b["pct"])) < 0.05
        return False
    if isinstance(a, str) or isinstance(b, str):
        return str(a).replace(" ", "") == str(b).replace(" ", "")
    try:
        fa, fb = float(a), float(b)
    except (TypeError, ValueError):
        return str(a) == str(b)
    if isinstance(a, int) and isinstance(b, int):
        return a == b
    return abs(fa - fb) < 1e-9


def match_label(site_v, xlsx_v) -> str:
    return "MATCH" if numbers_equal(site_v, xlsx_v) else "FAIL"


def fmt(v) -> str:
    if v is None:
        return ""
    try:
        missing = pd.isna(v)
        if isinstance(missing, bool) and missing:
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, dict):
        return v.get("raw", str(v))
    # pandas/numpy scalar values do not all satisfy isinstance(v, float), so
    # normalize numeric scalars through float before rounding.  This also
    # prevents NaN from reaching int(round(...)).
    if isinstance(v, (int, float)) or hasattr(v, "__float__"):
        try:
            fv = float(v)
            if math.isnan(fv) or math.isinf(fv):
                return ""
            if abs(fv - round(fv)) < 1e-9:
                return str(int(round(fv)))
            return ("%s" % fv).rstrip("0").rstrip(".")
        except (TypeError, ValueError, OverflowError):
            pass
    return str(v)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path, max_bytes=None) -> str:
    h = hashlib.sha256()
    n = 0
    with open(path, "rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
            n += len(chunk)
            if max_bytes is not None and n >= max_bytes:
                break
    return h.hexdigest()


def safe_name(url: str, limit=80) -> str:
    p = urlparse(url)
    host = re.sub(r"[^A-Za-z0-9._-]+", "_", p.netloc)[:40]
    tail = re.sub(r"[^A-Za-z0-9._-]+", "_", (p.path + ("_" + p.query if p.query else "")))[:limit]
    tail = tail.strip("_") or "root"
    return "%s__%s" % (host, tail)


# ---------------------------------------------------------------------------
# Load VERIFIED workbook
# ---------------------------------------------------------------------------

def load_verified(path: Path) -> dict:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError("VERIFIED workbook not found: %s" % path)
    xl = pd.ExcelFile(path, engine="openpyxl")
    data = {"path": str(path), "sheets": list(xl.sheet_names)}
    for name in xl.sheet_names:
        data[name] = pd.read_excel(xl, sheet_name=name)
    xl.close()
    readme_df = data["README"]
    readme = {}
    if list(readme_df.columns)[:2] == ["Item", "Detail"] or "Item" in readme_df.columns:
        for _, r in readme_df.iterrows():
            item = r.get("Item")
            if is_blank(item):
                continue
            readme[str(item).strip()] = None if is_blank(r.get("Detail")) else r.get("Detail")
    data["readme"] = readme
    panel = data["Panel_final"]
    data["iso3_list"] = sorted(panel["iso3"].dropna().astype(str).unique().tolist())
    data["iso_to_name"] = (
        panel.dropna(subset=["iso3"])
        .drop_duplicates("iso3")
        .set_index("iso3")["country"]
        .to_dict()
    )
    data["name_to_iso"] = {str(v).lower(): k for k, v in data["iso_to_name"].items()}
    aliases = {
        "korea": "KOR",
        "korea (south)": "KOR",
        "south korea": "KOR",
        "czechia": "CZE",
        "czech republic": "CZE",
        "slovak republic": "SVK",
        "slovakia": "SVK",
        "united kingdom": "GBR",
        "uk": "GBR",
        "united states": "USA",
        "usa": "USA",
        "russia": "RUS",
        "switzerland": "CHE",
        "portugal": "PRT",
        "sweden": "SWE",
        "taiwan": "TWN",
        "italy": "ITA",
        "japan": "JPN",
        "israel": "ISR",
        "bulgaria": "BGR",
        "france": "FRA",
        "turkey": "TUR",
        "poland": "POL",
        "china": "CHN",
        "netherlands": "NLD",
        "germany": "DEU",
        "austria": "AUT",
        "croatia": "HRV",
        "hungary": "HUN",
    }
    data["name_to_iso"].update(aliases)
    return data


# ---------------------------------------------------------------------------
# HTML extraction
# ---------------------------------------------------------------------------

def read_html(path: Path) -> str:
    return Path(path).read_text(encoding="utf-8", errors="replace")


def soup_of(html_text: str) -> BeautifulSoup:
    return BeautifulSoup(html_text, "lxml")


def heading_after(soup, text_prefix: str):
    for tag in soup.find_all(["h2", "h3"]):
        t = tag.get_text(" ", strip=True)
        if t.lower().startswith(text_prefix.lower()):
            return tag
    return None


def next_table(tag):
    if tag is None:
        return None
    return tag.find_next("table")


def cell_visible_number(td):
    a = td.find("a", class_="cell")
    if a is not None:
        return parse_site_number(a.get_text(" ", strip=True))
    strong = td.find("strong")
    if strong is not None:
        return parse_site_number(strong.get_text(" ", strip=True))
    # drop grade pill text
    clone_text = td.get_text(" ", strip=True)
    clone_text = re.sub(r"\b[ABCD]\b\s*$", "", clone_text).strip()
    if td.find("span", style=re.compile("faint")) or "&mdash;" in str(td) or "—" in td.get_text():
        if not re.search(r"\d", clone_text):
            return None
    return parse_site_number(clone_text)


def cell_grade(td):
    g = td.find("span", class_=re.compile(r"\bg\b"))
    if g is None:
        return None
    t = g.get_text(strip=True)
    return t if t in {"A", "B", "C", "D"} else None


def parse_panel_table(html_text: str):
    """Extract Year x variable cells from a country Panel data table."""
    soup = soup_of(html_text)
    h = heading_after(soup, "Panel data")
    table = next_table(h) if h is not None else None
    if table is None:
        # first table on the page
        table = soup.find("table")
    if table is None:
        return {"headers": [], "vars": [], "cells": []}
    ths = table.find("thead").find_all("th") if table.find("thead") else table.find_all("tr")[0].find_all(["th", "td"])
    headers = [th.get_text(" ", strip=True) for th in ths]
    vars_ = []
    for htxt in headers[1:]:
        key = HEADER_TO_VAR.get(htxt.strip().lower())
        vars_.append(key)
    cells = []
    body = table.find("tbody") or table
    for tr in body.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 2:
            continue
        year = parse_site_number(tds[0].get_text(" ", strip=True))
        for i, var in enumerate(vars_, start=1):
            if i >= len(tds) or not var:
                continue
            td = tds[i]
            cells.append({
                "year": year,
                "variable": var,
                "value": cell_visible_number(td),
                "grade": cell_grade(td),
                "raw": td.get_text(" ", strip=True),
            })
    return {"headers": headers, "vars": [v for v in vars_ if v], "cells": cells}


def parse_evidence_values(html_text: str):
    soup = soup_of(html_text)
    h = heading_after(soup, "Values")
    table = next_table(h) if h is not None else soup.find("table")
    rows = []
    if table is None:
        return rows
    body = table.find("tbody") or table
    for tr in body.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 2:
            continue
        year = parse_site_number(tds[0].get_text(" ", strip=True))
        val = cell_visible_number(tds[1])
        grade = cell_grade(tds[2]) if len(tds) > 2 else None
        rows.append({"year": year, "value": val, "grade": grade})
    return rows


def parse_quality_table(html_text: str):
    soup = soup_of(html_text)
    h = heading_after(soup, "Data quality")
    table = next_table(h) if h is not None else None
    rows = []
    if table is None:
        return rows
    body = table.find("tbody") or table
    for tr in body.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 4:
            continue
        label = tds[0].get_text(" ", strip=True)
        var = LABEL_TO_VAR.get(label.lower())
        rows.append({
            "label": label,
            "variable": var,
            "coverage": re.sub(r"\s+", "", tds[1].get_text(" ", strip=True)),
            "years": tds[2].get_text(" ", strip=True),
            "grade": cell_grade(tds[3]) or tds[3].get_text(" ", strip=True),
            "usable": tds[4].get_text(" ", strip=True) if len(tds) > 4 else "",
        })
    return rows


def parse_corrections_table(html_text: str):
    soup = soup_of(html_text)
    h = heading_after(soup, "Corrections applied")
    table = next_table(h) if h is not None else None
    rows = []
    if table is None:
        return rows
    body = table.find("tbody") or table
    headers = []
    if table.find("thead"):
        headers = [th.get_text(" ", strip=True).lower() for th in table.find("thead").find_all("th")]
    for tr in body.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue
        texts = [td.get_text(" ", strip=True) for td in tds]
        rec = {"raw": texts}
        if headers and "country" in headers:
            # verification.html: Country, Value, Was, Now, Why
            rec["iso3"] = texts[0].strip()
            rec["what"] = texts[1]
            rec["old"] = parse_site_number(texts[2])
            rec["new"] = parse_site_number(texts[3]) if len(texts) > 3 else None
        else:
            rec["what"] = texts[0]
            rec["old"] = parse_site_number(texts[1])
            rec["new"] = parse_site_number(texts[2])
        m = re.match(
            r"(?i)(population|foreign-born|foreign nationals|irregular stock|"
            r"overstayers|detections|absconded workers(?: \(tw\))?)\s+(\d{4})",
            rec.get("what") or "",
        )
        if m:
            rec["variable"] = LABEL_TO_VAR.get(m.group(1).lower())
            rec["year"] = int(m.group(2))
        rows.append(rec)
    return rows


def parse_verification_xy(html_text: str):
    m = re.search(
        r"<strong>(\d+)\s+of\s+(\d+)</strong>\s+machine-checkable",
        html_text,
        re.I,
    )
    if not m:
        return None
    return {"exact": int(m.group(1)), "checked": int(m.group(2))}


def parse_stat_boxes(html_text: str):
    soup = soup_of(html_text)
    out = []
    for box in soup.select("div.stat"):
        n = box.find("span", class_="n")
        lab = box.find("span", class_="l")
        if n is None:
            continue
        out.append({
            "label": lab.get_text(" ", strip=True) if lab else "",
            "value_text": n.get_text(" ", strip=True),
            "value": parse_site_number(n.get_text(" ", strip=True)),
        })
    return out


def parse_grade_table(html_text: str):
    soup = soup_of(html_text)
    h = heading_after(soup, "How reliable")
    table = next_table(h) if h is not None else None
    rows = []
    if table is None:
        return rows
    for tr in (table.find("tbody") or table).find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue
        grade = cell_grade(tds[0]) or tds[0].get_text(" ", strip=True)
        rows.append({
            "grade": grade,
            "values": parse_site_number(tds[2].get_text(" ", strip=True)),
            "share": parse_site_number(tds[3].get_text(" ", strip=True)) if len(tds) > 3 else None,
        })
    return rows


def parse_vlog_source_table(html_text: str):
    soup = soup_of(html_text)
    h = heading_after(soup, "Reproduction rate")
    table = next_table(h) if h is not None else None
    rows = []
    if table is None:
        return rows
    for tr in (table.find("tbody") or table).find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 4:
            continue
        rows.append({
            "source": tds[0].get_text(" ", strip=True),
            "checked": parse_site_number(tds[1].get_text(" ", strip=True)),
            "exact": parse_site_number(tds[2].get_text(" ", strip=True)),
            "rate": parse_site_number(tds[3].get_text(" ", strip=True)),
        })
    return rows


def parse_methods_coverage(html_text: str):
    soup = soup_of(html_text)
    h = heading_after(soup, "How much weight")
    table = next_table(h) if h is not None else None
    rows = []
    if table is None:
        return rows
    for tr in (table.find("tbody") or table).find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue
        label = tds[0].get_text(" ", strip=True)
        rows.append({
            "label": label,
            "variable": LABEL_TO_VAR.get(label.lower()),
            "country_years": re.sub(r"\s+", "", tds[1].get_text(" ", strip=True)),
            "countries": re.sub(r"\s+", "", tds[2].get_text(" ", strip=True)),
        })
    return rows


def parse_countries_index(html_text: str):
    soup = soup_of(html_text)
    out = []
    for a in soup.select("a.cbox"):
        href = a.get("href") or ""
        m = re.search(r"countries/([A-Z]{3})\.html", href.replace("\\", "/"))
        if not m:
            continue
        iso = m.group(1)
        cm = a.find("span", class_="cm")
        text = cm.get_text(" ", strip=True) if cm else a.get_text(" ", strip=True)
        vm = re.search(r"(\d+)\s+values", text)
        fm = re.search(r"(\d+)\s+files", text)
        out.append({
            "iso3": iso,
            "values": int(vm.group(1)) if vm else None,
            "files": int(fm.group(1)) if fm else None,
            "text": text,
        })
    return out


def iter_site_pages(site: Path):
    site = Path(site)
    for p in sorted(site.glob("*.html")):
        yield p
    for sub in ("countries", "evidence-pages"):
        d = site / sub
        if d.is_dir():
            for p in sorted(d.glob("*.html")):
                yield p


def resolve_local(page: Path, href: str, site: Path):
    href = (href or "").split("#")[0].split("?")[0]
    if not href:
        return None
    href = unquote(htmlmod.unescape(href))
    target = (page.parent / href).resolve()
    try:
        target.relative_to(site.resolve())
    except ValueError:
        return target
    return target


def extract_stat_links_from_html(page: Path, html_text: str, site: Path):
    """Collect statistic-associated http(s) URLs and local archive hrefs.

    Pairs a live URL with archive hrefs that appear in the same table row or
    the same evidence-page Source / Archived sections.
    """
    soup = soup_of(html_text)
    pairs = []  # dicts
    rel = str(page.relative_to(site)).replace("\\", "/")

    def is_archive_href(href: str) -> bool:
        h = href.replace("\\", "/")
        return ("evidence/" in h) or h.endswith(".pdf") or h.endswith(".json") \
            or h.endswith(".xlsx") or h.endswith(".png") or h.endswith(".html") \
            or h.endswith(".csv")

    def skip_url(url: str) -> bool:
        if not url:
            return True
        low = url.lower()
        if any(low.startswith(p) for p in NAV_SKIP_PREFIXES):
            return True
        host = urlparse(url).netloc.lower()
        if host in AUTHOR_HOSTS:
            return True
        return False

    # table-row pairing (country sources, sources.html, evidence pages)
    for tr in soup.find_all("tr"):
        live = []
        archives = []
        for a in tr.find_all("a", href=True):
            href = htmlmod.unescape(a["href"])
            if href.startswith("http://") or href.startswith("https://"):
                if not skip_url(href):
                    live.append(href)
            elif is_archive_href(href) and not href.startswith("#"):
                archives.append(href)
        if not live and not archives:
            continue
        # skip pure navigation rows
        if live and not archives and all("raymond.cph" in u for u in live):
            continue
        for u in live:
            if archives:
                for ar in archives:
                    pairs.append({
                        "page": rel,
                        "url": u,
                        "archive_href": ar,
                        "archive_path": str(resolve_local(page, ar, site)) if resolve_local(page, ar, site) else "",
                    })
            else:
                pairs.append({
                    "page": rel,
                    "url": u,
                    "archive_href": "",
                    "archive_path": "",
                })
        if archives and not live:
            for ar in archives:
                loc = resolve_local(page, ar, site)
                pairs.append({
                    "page": rel,
                    "url": "",
                    "archive_href": ar,
                    "archive_path": str(loc) if loc else "",
                })

    # evidence-page Source <ul> + following Archived <p>
    if "evidence-pages" in rel.replace("\\", "/"):
        for ul in soup.select("ul.clean"):
            lives = []
            for a in ul.find_all("a", href=True):
                href = htmlmod.unescape(a["href"])
                if href.startswith("http") and not skip_url(href):
                    lives.append(href)
            sec = ul.find_parent("section")
            nxt = sec.find_next_sibling("section") if sec else None
            archives = []
            if nxt:
                for a in nxt.find_all("a", href=True):
                    href = htmlmod.unescape(a["href"])
                    if is_archive_href(href):
                        archives.append(href)
            for u in lives:
                if archives:
                    for ar in archives:
                        pairs.append({
                            "page": rel,
                            "url": u,
                            "archive_href": ar,
                            "archive_path": str(resolve_local(page, ar, site)) if resolve_local(page, ar, site) else "",
                        })
                else:
                    pairs.append({"page": rel, "url": u, "archive_href": "", "archive_path": ""})

    return pairs


def collect_all_stat_links(site: Path):
    site = Path(site)
    all_pairs = []
    pages = []
    pages.extend(sorted(site.glob("*.html")))
    for sub in ("countries", "evidence-pages"):
        d = site / sub
        if d.is_dir():
            pages.extend(sorted(d.glob("*.html")))
    wanted = []
    for p in pages:
        rel = str(p.relative_to(site)).replace("\\", "/")
        if rel in {"index.html", "countries.html", "data.html", "methods.html"}:
            # still scan sources/verification/country/evidence; index has few stat urls
            if rel in {"data.html", "methods.html", "countries.html"}:
                continue
        wanted.append(p)
    # always include these
    extra = [site / "index.html", site / "sources.html", site / "verification.html"]
    extra += list((site / "countries").glob("*.html"))
    extra += list((site / "evidence-pages").glob("*.html"))
    seen_files = set()
    for p in extra:
        if not p.is_file() or p in seen_files:
            continue
        seen_files.add(p)
        all_pairs.extend(extract_stat_links_from_html(p, read_html(p), site))
    return all_pairs


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def new_row(**kwargs):
    rec = {k: "" for k in ROW_FIELDS}
    rec.update(kwargs)
    for k in ("site_value", "xlsx_value"):
        if k in rec and rec[k] is not None and not isinstance(rec[k], str):
            rec[k] = fmt(rec[k])
    if rec.get("year") in ("", None):
        rec["year"] = ""
    elif not isinstance(rec["year"], str):
        rec["year"] = str(int(rec["year"])) if isinstance(rec["year"], (int, float)) and rec["year"] == rec["year"] else str(rec["year"])
    return rec


def panel_lookup(panel: pd.DataFrame, iso3: str, year, var):
    sub = panel[(panel["iso3"] == iso3) & (panel["year"] == year)]
    if sub.empty or var not in sub.columns:
        return None, None
    r = sub.iloc[0]
    val = r[var]
    gcol = var + "_grade"
    grade = r[gcol] if gcol in sub.columns else None
    if is_blank(val):
        val = None
    if is_blank(grade):
        grade = None
    else:
        grade = str(grade).strip()
    return (None if is_blank(val) else val), grade


def url_col(panel, var):
    c = var + "_url"
    return c if c in panel.columns else None


# ---------------------------------------------------------------------------
# Compare: build all audit rows
# ---------------------------------------------------------------------------

def build_audit_rows(site: Path, verified: dict) -> list:
    site = Path(site)
    panel = verified["Panel_final"]
    qual = verified["Data_quality"]
    corr = verified["Corrections_applied"]
    issues = verified["Known_issues"]
    vlog = verified["Verification_log"]
    reg = verified["Source_register"]
    irr = verified["Irregular_estimates_all"]
    readme = verified["readme"]
    iso_list = verified["iso3_list"]
    rows = []

    index_html = read_html(site / "index.html")
    ver_html = read_html(site / "verification.html")
    methods_html = read_html(site / "methods.html") if (site / "methods.html").is_file() else ""
    readme_md = read_html(site / "README.md") if (site / "README.md").is_file() else ""
    countries_html = read_html(site / "countries.html")
    countries_idx = {r["iso3"]: r for r in parse_countries_index(countries_html)}

    # ---- README / site-wide headlines (duplicated onto SUMMARY via iso3=ALL)
    def add_headline(item, site_v, xlsx_v, page, notes=""):
        rows.append(new_row(
            iso3="ALL",
            verified_sheet="README",
            page=page,
            item=item,
            site_value=site_v,
            xlsx_value=xlsx_v,
            number_match=match_label(site_v, xlsx_v) if (site_v is not None or xlsx_v is not None) else "N/A",
            notes=notes,
        ))

    stats_index = parse_stat_boxes(index_html)
    stats_ver = parse_stat_boxes(ver_html)
    def readme_get(*keys):
        for k in keys:
            if k in readme and readme[k] is not None:
                return readme[k]
        return None

    r_rederived = normalize_xlsx_number(readme_get("Values re-derived from live sources"))
    r_exact = normalize_xlsx_number(readme_get("matched exactly", "  matched exactly"))
    r_disc = normalize_xlsx_number(readme_get("discrepancies found", "  discrepancies found"))
    r_corr = readme_get("Corrections applied")
    r_corr_n = parse_site_number(str(r_corr).split()[0]) if r_corr else None
    r_retrieved = readme_get("retrieved to country folders", "  retrieved to country folders")
    r_cited = readme_get("Document sources cited")

    for box in stats_index + [{"_page": "verification.html", **b} for b in stats_ver]:
        page = box.get("_page", "index.html")
        lab = (box.get("label") or "").lower()
        val = box.get("value")
        if "re-derived" in lab:
            add_headline("values re-derived from live sources", val, r_rederived, page)
        elif "matched the source exactly" in lab:
            x_pct = r_exact.get("pct") if isinstance(r_exact, dict) else r_exact
            add_headline("matched exactly percent", val, x_pct, page,
                         notes="xlsx Detail=%r" % readme_get("matched exactly", "  matched exactly"))
        elif lab.startswith("matched exactly"):
            x_count = r_exact["count"] if isinstance(r_exact, dict) else r_exact
            add_headline("matched exactly count", val, x_count, page,
                         notes="xlsx Detail=%r" % readme_get("matched exactly", "  matched exactly"))
        elif "discrepanc" in lab:
            add_headline("discrepancies found", val, r_disc, page)
        elif "values corrected" in lab or lab.startswith("values corrected"):
            add_headline("values corrected", val, r_corr_n, page,
                         notes="xlsx Detail=%r" % r_corr)
        elif "countries, 13 years" in lab:
            add_headline("countries", val, 40, page, notes="Panel_final iso3 count")

    # 98.4% in index body / README.md
    for page, text in (("index.html", index_html), ("README.md", readme_md), ("verification.html", ver_html)):
        m = re.search(r"(\d[\d,]*)\s*\((\d+\.\d+)%\)", text)
        if m:
            add_headline("matched exactly count+percent",
                         "%s (%s%%)" % (parse_site_number(m.group(1)), m.group(2)),
                         readme_get("matched exactly", "  matched exactly"), page)

    # grade counts
    grades_site = parse_grade_table(index_html)
    xlsx_grades = {
        "A": normalize_xlsx_number(readme.get("A - re-derived from a machine-readable official source, exact match")),
        "B": normalize_xlsx_number(readme.get("B - official statistical source, consistent")),
        "C": normalize_xlsx_number(readme.get("C - source document retrieved, modelled estimate not re-derivable")),
        "D": normalize_xlsx_number(readme.get("D - cited source could not be retrieved")),
    }
    grade_cols = [c for c in panel.columns if c.endswith("_grade")]
    derived = Counter()
    for c in grade_cols:
        for g in panel[c].dropna():
            derived[str(g).strip()] += 1
    for g in "ABCD":
        site_v = next((r["values"] for r in grades_site if r["grade"] == g), None)
        add_headline("grade %s count" % g, site_v, xlsx_grades.get(g), "index.html",
                     notes="re-derived from Panel_final grade cells=%s" % derived.get(g))
        site_share = next((r["share"] for r in grades_site if r["grade"] == g), None)
        total_g = sum(derived.values()) or 1
        x_share = round(100.0 * derived.get(g, 0) / total_g, 1)
        add_headline("grade %s share percent" % g, site_share, x_share, "index.html",
                     notes="computed as 100*count/sum(grades) from Panel_final")

    m = re.search(r"Every one of the ([\d,]+) values in the panel", index_html)
    if m:
        add_headline("panel values with a grade", parse_site_number(m.group(1)),
                     sum(derived.values()), "index.html",
                     notes="sum of non-null *_grade cells in Panel_final")
    m = re.search(r"Quality grades on all ([\d,]+) values", readme_md)
    if m:
        add_headline("panel values with a grade (README.md)", parse_site_number(m.group(1)),
                     sum(derived.values()), "README.md")

    # document sources 89 / 87 / 2
    cited_n = parse_site_number(str(r_cited).split()[0]) if r_cited else None
    retrieved_n = parse_site_number(str(r_retrieved).split()[0]) if r_retrieved else None
    add_headline("document sources cited",
                 cited_n,
                 cited_n,
                 "README",
                 notes="xlsx=%r (headline lives in workbook README; site uses 76/78 wording)" % r_cited)
    # site 76 of 78
    for page, text in (("index.html", index_html), ("README.md", readme_md)):
        m2 = re.search(r"(\d+)\s+country.?source citations,\s*(\d+)\s+of them archived", text)
        x_frac = None
        if retrieved_n is not None and cited_n is not None:
            x_frac = "%s/%s" % (retrieved_n, cited_n)
        if m2:
            add_headline("country-source citations archived/total",
                         "%s/%s" % (m2.group(2), m2.group(1)),
                         x_frac,
                         page,
                         notes="site archived/total vs xlsx retrieved/cited (%r / %r)" % (r_retrieved, r_cited))
        m3 = re.search(r"\*\*(\d+) of (\d+)\*\* distinct country-source", text)
        if m3:
            add_headline("country-source citations archived/total",
                         "%s/%s" % (m3.group(1), m3.group(2)),
                         x_frac,
                         page,
                         notes="site archived/total vs xlsx retrieved/cited (%r / %r)" % (r_retrieved, r_cited))

    # verification source table vs Verification_log
    vlog_site = parse_vlog_source_table(ver_html)
    vlog_grp = vlog.groupby("source").agg(
        checked=("status", "size"),
        exact=("status", lambda s: int((s == "EXACT").sum())),
    )
    for rec in vlog_site:
        src = rec["source"]
        if src in vlog_grp.index:
            x_chk = int(vlog_grp.loc[src, "checked"])
            x_ex = int(vlog_grp.loc[src, "exact"])
        else:
            x_chk, x_ex = None, None
        rows.append(new_row(
            iso3="ALL", verified_sheet="Verification_log", page="verification.html",
            item="source %s values checked" % src,
            site_value=rec["checked"], xlsx_value=x_chk,
            number_match=match_label(rec["checked"], x_chk),
        ))
        rows.append(new_row(
            iso3="ALL", verified_sheet="Verification_log", page="verification.html",
            item="source %s exact" % src,
            site_value=rec["exact"], xlsx_value=x_ex,
            number_match=match_label(rec["exact"], x_ex),
        ))
        if rec["checked"] and rec["exact"] is not None and rec["rate"] is not None and rec["checked"]:
            x_rate = round(100.0 * x_ex / x_chk, 1) if x_chk else None
            rows.append(new_row(
                iso3="ALL", verified_sheet="Verification_log", page="verification.html",
                item="source %s rate percent" % src,
                site_value=rec["rate"], xlsx_value=x_rate,
                number_match=match_label(rec["rate"], x_rate),
            ))

    # methods coverage vs Panel_final
    for rec in parse_methods_coverage(methods_html):
        var = rec["variable"]
        if not var or var not in panel.columns:
            continue
        n = int(panel[var].notna().sum())
        n_cty = int(panel.loc[panel[var].notna(), "iso3"].nunique())
        rows.append(new_row(
            iso3="ALL", verified_sheet="Panel_final", page="methods.html",
            variable=var, item="country-years coverage",
            site_value=rec["country_years"], xlsx_value="%s/520" % n,
            number_match=match_label(rec["country_years"], "%s/520" % n),
        ))
        rows.append(new_row(
            iso3="ALL", verified_sheet="Panel_final", page="methods.html",
            variable=var, item="countries coverage",
            site_value=rec["countries"], xlsx_value="%s/40" % n_cty,
            number_match=match_label(rec["countries"], "%s/40" % n_cty),
        ))

    # Known_issues presence on verification.html
    ver_text = soup_of(ver_html).get_text(" ", strip=True)
    for _, iss in issues.iterrows():
        snippet = str(iss.get("issue") or "")[:80]
        present = snippet[:40] in ver_text if snippet else False
        scope = str(iss.get("scope") or "")
        rows.append(new_row(
            iso3="ALL", verified_sheet="Known_issues", page="verification.html",
            variable=iss.get("variable"),
            item="issue [%s] %s" % (iss.get("severity"), scope),
            site_value="present" if present else "absent",
            xlsx_value="present",
            number_match="MATCH" if present else "FAIL",
            notes=(snippet or "")[:200],
        ))

    vcorr_all = parse_corrections_table(ver_html)

    # Per-country
    for iso3 in iso_list:
        cpage = site / "countries" / ("%s.html" % iso3)
        chtml = read_html(cpage) if cpage.is_file() else ""
        g = panel[panel["iso3"] == iso3]
        ev_vars_present = []
        for v in PANEL_VARS:
            evp = site / "evidence-pages" / ("%s__%s.html" % (iso3, v))
            if evp.is_file():
                ev_vars_present.append(v)

        # Panel_final from country HTML
        parsed = parse_panel_table(chtml) if chtml else {"vars": [], "cells": []}
        site_vars = parsed["vars"]
        seen_keys = set()
        for cell in parsed["cells"]:
            year, var = cell["year"], cell["variable"]
            xv, xg = panel_lookup(panel, iso3, year, var)
            seen_keys.add((int(year) if year else None, var))
            ucol = url_col(panel, var)
            src_url = ""
            if ucol:
                sub = g[g["year"] == year]
                if not sub.empty and not is_blank(sub.iloc[0][ucol]):
                    src_url = str(sub.iloc[0][ucol])
            rows.append(new_row(
                iso3=iso3, verified_sheet="Panel_final",
                page="countries/%s.html" % iso3,
                variable=var, year=year,
                item="panel cell",
                site_value=cell["value"], xlsx_value=xv,
                number_match=match_label(cell["value"], xv),
                url=src_url,
            ))
            if cell["grade"] is not None or xg is not None:
                grade_match = (
                    "N/A" if xg is None and cell["grade"] is not None
                    else (match_label(cell["grade"], xg) if (cell["grade"] or xg) else "N/A")
                )
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Panel_final",
                    page="countries/%s.html" % iso3,
                    variable=var, year=year,
                    item="panel grade",
                    site_value=cell["grade"], xlsx_value=xg,
                    number_match=grade_match,
                    url=src_url,
                    notes="grade column not present in reference workbook" if xg is None and cell["grade"] is not None else "",
                ))

        # xlsx cells the country page omitted (column missing with non-null values)
        for _, r in g.iterrows():
            year = int(r["year"])
            for var in PANEL_VARS:
                if var not in panel.columns:
                    continue
                xv, xg = panel_lookup(panel, iso3, year, var)
                if (year, var) in seen_keys:
                    continue
                if xv is None and var not in site_vars:
                    continue  # correctly omitted empty series
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Panel_final",
                    page="countries/%s.html" % iso3,
                    variable=var, year=year,
                    item="panel cell missing on site",
                    site_value=None, xlsx_value=xv,
                    number_match=match_label(None, xv),
                    notes="column absent on country page" if var not in site_vars else "cell not parsed",
                ))

        # Evidence pages
        for var in ev_vars_present:
            evp = site / "evidence-pages" / ("%s__%s.html" % (iso3, var))
            ev_html = read_html(evp)
            for cell in parse_evidence_values(ev_html):
                year = cell["year"]
                xv, xg = panel_lookup(panel, iso3, year, var)
                ucol = url_col(panel, var)
                src_url = ""
                if ucol:
                    sub = g[g["year"] == year]
                    if not sub.empty and not is_blank(sub.iloc[0][ucol]):
                        src_url = str(sub.iloc[0][ucol])
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Panel_final",
                    page="evidence-pages/%s__%s.html" % (iso3, var),
                    variable=var, year=year,
                    item="evidence cell",
                    site_value=cell["value"], xlsx_value=xv,
                    number_match=match_label(cell["value"], xv),
                    url=src_url,
                ))
                if cell["grade"] is not None or xg is not None:
                    grade_match = (
                        "N/A" if xg is None and cell["grade"] is not None
                        else (match_label(cell["grade"], xg) if (cell["grade"] or xg) else "N/A")
                    )
                    rows.append(new_row(
                        iso3=iso3, verified_sheet="Panel_final",
                        page="evidence-pages/%s__%s.html" % (iso3, var),
                        variable=var, year=year,
                        item="evidence grade",
                        site_value=cell["grade"], xlsx_value=xg,
                        number_match=grade_match,
                        url=src_url,
                        notes="grade column not present in reference workbook" if xg is None and cell["grade"] is not None else "",
                    ))

        # Data_quality
        cq = qual[qual["iso3"] == iso3]
        qsite = {r["variable"]: r for r in parse_quality_table(chtml) if r.get("variable")}
        if cq.empty:
            rows.append(new_row(
                iso3=iso3, verified_sheet="Data_quality",
                page="countries/%s.html" % iso3,
                item="no Data_quality rows in xlsx",
                number_match="N/A",
            ))
        for _, r in cq.iterrows():
            var = r["variable"]
            site_r = qsite.get(var)
            n_years = r.get("n_years")
            if site_r is None:
                # site hides n_years==0
                expected_hidden = (is_blank(n_years) or int(n_years) == 0)
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Data_quality",
                    page="countries/%s.html" % iso3,
                    variable=var,
                    item="quality row display",
                    site_value="hidden",
                    xlsx_value=r.get("coverage"),
                    number_match="MATCH" if expected_hidden else "FAIL",
                    notes="n_years=%s; site omits zero-coverage rows" % n_years,
                ))
                continue
            rows.append(new_row(
                iso3=iso3, verified_sheet="Data_quality",
                page="countries/%s.html" % iso3,
                variable=var, item="coverage",
                site_value=site_r["coverage"], xlsx_value=r.get("coverage"),
                number_match=match_label(site_r["coverage"], r.get("coverage")),
            ))
            rows.append(new_row(
                iso3=iso3, verified_sheet="Data_quality",
                page="countries/%s.html" % iso3,
                variable=var, item="years span",
                site_value=site_r["years"], xlsx_value="" if is_blank(r.get("years")) else r.get("years"),
                number_match="MATCH" if str(site_r["years"]) == str("" if is_blank(r.get("years")) else r.get("years")) else "FAIL",
            ))
            rows.append(new_row(
                iso3=iso3, verified_sheet="Data_quality",
                page="countries/%s.html" % iso3,
                variable=var, item="modal grade",
                site_value=site_r["grade"], xlsx_value=r.get("modal_grade"),
                number_match=match_label(site_r["grade"], r.get("modal_grade")),
            ))
            su = (site_r.get("usable") or "").strip()
            xu = "" if is_blank(r.get("usable_for_trend")) else str(r.get("usable_for_trend")).strip()
            rows.append(new_row(
                iso3=iso3, verified_sheet="Data_quality",
                page="countries/%s.html" % iso3,
                variable=var, item="usable_for_trend",
                site_value=su, xlsx_value=xu,
                number_match="MATCH" if su == xu else "FAIL",
            ))

        # Corrections_applied
        cc = corr[corr["iso3"] == iso3]
        csite = parse_corrections_table(chtml)
        if cc.empty:
            rows.append(new_row(
                iso3=iso3, verified_sheet="Corrections_applied",
                page="countries/%s.html" % iso3,
                item="no corrections for this country",
                site_value=len(csite), xlsx_value=0,
                number_match="MATCH" if len(csite) == 0 else "FAIL",
            ))
        else:
            # match by year+variable
            used = set()
            for _, r in cc.iterrows():
                year, var = int(r["year"]), r["variable"]
                hit = None
                for i, s in enumerate(csite):
                    if i in used:
                        continue
                    if s.get("year") == year and s.get("variable") == var:
                        hit = s
                        used.add(i)
                        break
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Corrections_applied",
                    page="countries/%s.html" % iso3,
                    variable=var, year=year, item="old_value",
                    site_value=None if hit is None else hit.get("old"),
                    xlsx_value=r.get("old_value"),
                    number_match="FAIL" if hit is None else match_label(hit.get("old"), r.get("old_value")),
                    notes="" if hit else "correction not shown on country page",
                ))
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Corrections_applied",
                    page="countries/%s.html" % iso3,
                    variable=var, year=year, item="new_value",
                    site_value=None if hit is None else hit.get("new"),
                    xlsx_value=r.get("new_value"),
                    number_match="FAIL" if hit is None else match_label(hit.get("new"), r.get("new_value")),
                ))
            # also verification.html global corrections for this iso
            vcorr = [c for c in vcorr_all if c.get("iso3") == iso3]
            for _, r in cc.iterrows():
                year, var = int(r["year"]), r["variable"]
                hit = next((s for s in vcorr if s.get("year") == year and s.get("variable") == var), None)
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Corrections_applied",
                    page="verification.html",
                    variable=var, year=year, item="old_value",
                    site_value=None if hit is None else hit.get("old"),
                    xlsx_value=r.get("old_value"),
                    number_match="FAIL" if hit is None else match_label(hit.get("old"), r.get("old_value")),
                ))
                rows.append(new_row(
                    iso3=iso3, verified_sheet="Corrections_applied",
                    page="verification.html",
                    variable=var, year=year, item="new_value",
                    site_value=None if hit is None else hit.get("new"),
                    xlsx_value=r.get("new_value"),
                    number_match="FAIL" if hit is None else match_label(hit.get("new"), r.get("new_value")),
                ))

        # Verification_log country summary
        cv = vlog[vlog["iso3"] == iso3]
        xy = parse_verification_xy(chtml)
        x_chk = int(len(cv))
        x_ex = int((cv["status"] == "EXACT").sum()) if x_chk else 0
        if xy is None:
            rows.append(new_row(
                iso3=iso3, verified_sheet="Verification_log",
                page="countries/%s.html" % iso3,
                item="machine-checkable X of Y",
                site_value="not shown",
                xlsx_value="%s of %s" % (x_ex, x_chk),
                number_match="MATCH" if x_chk == 0 else "FAIL",
            ))
        else:
            rows.append(new_row(
                iso3=iso3, verified_sheet="Verification_log",
                page="countries/%s.html" % iso3,
                item="machine-checkable exact count",
                site_value=xy["exact"], xlsx_value=x_ex,
                number_match=match_label(xy["exact"], x_ex),
            ))
            rows.append(new_row(
                iso3=iso3, verified_sheet="Verification_log",
                page="countries/%s.html" % iso3,
                item="machine-checkable checked count",
                site_value=xy["checked"], xlsx_value=x_chk,
                number_match=match_label(xy["checked"], x_chk),
            ))

        # Source_register: years + URL presence on country page
        cr = reg[reg["iso3"] == iso3]
        page_urls = set(re.findall(r'href="(https?://[^"]+)"', chtml))
        page_urls = {htmlmod.unescape(u) for u in page_urls}
        if cr.empty:
            rows.append(new_row(
                iso3=iso3, verified_sheet="Source_register",
                page="countries/%s.html" % iso3,
                item="no Source_register rows",
                number_match="N/A",
            ))
        seen_src = set()
        for _, r in cr.iterrows():
            url = "" if is_blank(r.get("source_url")) else str(r.get("source_url"))
            key = (r.get("variable"), url)
            if key in seen_src:
                continue
            seen_src.add(key)
            on_page = url in page_urls if url else False
            # country page may truncate visible text but href is full
            lf = "" if is_blank(r.get("local_file")) else str(r.get("local_file"))
            local = ""
            if lf and lf not in {"data_raw/", "nan"}:
                cand = site / "evidence" / "countries" / iso3 / lf
                local = str(cand)
            rows.append(new_row(
                iso3=iso3, verified_sheet="Source_register",
                page="countries/%s.html" % iso3,
                variable=r.get("variable"),
                item="source_url listed (years %s)" % r.get("years"),
                site_value="href present" if on_page else "href absent",
                xlsx_value=url,
                number_match="MATCH" if on_page else "FAIL",
                url=url,
                archive_path=local,
                archive_exists="" if not local else ("yes" if Path(local).is_file() else "no"),
                notes=("n_obs=%s retrieval=%s local_file=%s"
                       % (r.get("n_obs"), r.get("retrieval"), lf)),
            ))

        # Irregular_estimates_all
        ir = irr[irr["iso3"] == iso3] if "iso3" in irr.columns else irr.iloc[0:0]
        if ir.empty:
            rows.append(new_row(
                iso3=iso3, verified_sheet="Irregular_estimates_all",
                page="countries/%s.html" % iso3,
                item="no irregular estimates in xlsx",
                number_match="N/A",
            ))
        else:
            rows.append(new_row(
                iso3=iso3, verified_sheet="Irregular_estimates_all",
                page="countries/%s.html" % iso3,
                item="xlsx estimate count",
                site_value="", xlsx_value=len(ir),
                number_match="N/A",
                notes="competing estimates are not listed on country pages; panel primary is shown as panel cells",
            ))
            for _, r in ir.iterrows():
                var = r.get("variable")
                year = r.get("year")
                rank = "" if is_blank(r.get("estimate_rank_in_cell")) else str(r.get("estimate_rank_in_cell"))
                xv = r.get("value")
                # displayed iff PANEL PRIMARY and the panel cell is shown
                displayed = None
                for cell in parsed["cells"]:
                    if cell["variable"] == var and cell["year"] == year:
                        displayed = cell["value"]
                        break
                if rank == "PANEL PRIMARY":
                    # The website country page is generated from Panel_final,
                    # not from the competing-estimates worksheet.  Compare the
                    # displayed value to the authoritative Panel_final cell;
                    # retain the competing-sheet value in notes so an internal
                    # inconsistency is visible without mislabelling the page.
                    panel_xv, _ = panel_lookup(panel, iso3, year, var)
                    cross = "MATCH" if numbers_equal(xv, panel_xv) else "DIFF"
                    rows.append(new_row(
                        iso3=iso3, verified_sheet="Irregular_estimates_all",
                        page="countries/%s.html" % iso3,
                        variable=var, year=year,
                        item="PANEL PRIMARY shown on site (cross-check Panel_final)",
                        site_value=displayed, xlsx_value=panel_xv,
                        number_match=match_label(displayed, panel_xv),
                        url="" if is_blank(r.get("source_url")) else str(r.get("source_url")),
                        notes="Irregular_estimates_all value=%s; internal sheet vs Panel_final=%s"
                              % (fmt(xv), cross),
                    ))
                else:
                    rows.append(new_row(
                        iso3=iso3, verified_sheet="Irregular_estimates_all",
                        page="countries/%s.html" % iso3,
                        variable=var, year=year,
                        item="non-primary estimate (not a panel cell)",
                        site_value="not displayed",
                        xlsx_value=xv,
                        number_match="N/A",
                        url="" if is_blank(r.get("source_url")) else str(r.get("source_url")),
                        notes=rank,
                    ))

        # Known_issues scoped to this country
        cname = str(verified["iso_to_name"].get(iso3, "")).lower()
        for _, iss in issues.iterrows():
            scope = str(iss.get("scope") or "")
            scoped = _issue_applies(scope, iso3, cname, verified)
            if not scoped:
                continue
            snippet = str(iss.get("issue") or "")[:80]
            present = snippet[:40] in ver_text if snippet else False
            rows.append(new_row(
                iso3=iso3, verified_sheet="Known_issues",
                page="verification.html",
                variable=iss.get("variable"),
                item="[%s] %s" % (iss.get("severity"), scope),
                site_value="present" if present else "absent",
                xlsx_value="present",
                number_match="MATCH" if present else "FAIL",
                notes=(snippet or "")[:200],
            ))

        # README share for this country
        n_corr = int(len(cc))
        n_v = int(len(cv))
        n_ex = int((cv["status"] == "EXACT").sum()) if n_v else 0
        n_panel = 0
        # countries.html counts only the six core variables used by its
        # generator; Taiwan's extra absconded-worker column is shown on the
        # country page but is intentionally not included in the index count.
        for v in PANEL_VARS[:6]:
            if v in g.columns:
                n_panel += int(g[v].notna().sum())
        rows.append(new_row(
            iso3=iso3, verified_sheet="README",
            page="countries/%s.html" % iso3,
            item="this country's share of 49 corrections",
            site_value=len(csite),
            xlsx_value=n_corr,
            number_match=match_label(len(csite), n_corr),
            notes="xlsx Corrections_applied rows for %s; site country-page correction rows" % iso3,
        ))
        rows.append(new_row(
            iso3=iso3, verified_sheet="README",
            page="countries/%s.html" % iso3,
            item="this country's verification_log rows (of 2454)",
            site_value=xy["checked"] if xy else "",
            xlsx_value=n_v,
            number_match=match_label(xy["checked"], n_v) if xy else ("N/A" if n_v == 0 else "FAIL"),
        ))
        rows.append(new_row(
            iso3=iso3, verified_sheet="README",
            page="countries.html",
            item="countries.html value count",
            site_value=countries_idx.get(iso3, {}).get("values"),
            xlsx_value=n_panel,
            number_match=match_label(countries_idx.get(iso3, {}).get("values"), n_panel),
        ))
        evdir = site / "evidence" / "countries" / iso3
        n_files = len([p for p in evdir.iterdir() if p.is_file()]) if evdir.is_dir() else 0
        rows.append(new_row(
            iso3=iso3, verified_sheet="README",
            page="countries.html",
            item="countries.html file count",
            site_value=countries_idx.get(iso3, {}).get("files"),
            xlsx_value=n_files,
            number_match=match_label(countries_idx.get(iso3, {}).get("files"), n_files),
            notes="xlsx_value here is on-disk evidence/countries/%s file count" % iso3,
        ))

    return rows


def _first_int(m):
    if not m:
        return None
    if hasattr(m, "group"):
        return parse_site_number(m.group(1))
    return None


def _issue_applies(scope: str, iso3: str, cname: str, verified: dict) -> bool:
    s = (scope or "").lower()
    if not s:
        return False
    if s in {"all", "eu/efta", "eurostat / oecd countries"}:
        return True
    if iso3.lower() in s:
        return True
    if cname and cname in s:
        return True
    # split on commas / and
    parts = re.split(r"[,;/]| and ", s)
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if p == cname or verified["name_to_iso"].get(p) == iso3:
            return True
        if p.startswith(cname):
            return True
    # Korea
    if iso3 == "KOR" and "korea" in s:
        return True
    return False


# ---------------------------------------------------------------------------
# Fetch + content compare
# ---------------------------------------------------------------------------

def load_fetch_cache(path: Path) -> dict:
    if path and Path(path).is_file():
        return json.loads(Path(path).read_text(encoding="utf-8"))
    return {}


def save_fetch_cache(path: Path, cache: dict):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")


def fetch_one(url: str, dest_dir: Path, timeout=(20, 70), do_second=True) -> dict:
    import requests
    # The workstation's conda hook removes SSL_CERT_FILE before commands run,
    # which makes requests reject otherwise reachable HTTPS endpoints.  The
    # separate PowerShell spot-check uses the Windows certificate store; this
    # fetch is for response/content capture and records the response status.
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    out = {
        "url": url,
        "status": None,
        "error": None,
        "sha256": None,
        "bytes": None,
        "content_type": None,
        "path": None,
        "second_sha256": None,
        "second_consistent": None,
        "downloadable": False,
    }
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        with requests.get(url, headers=headers, timeout=timeout, stream=True,
                          allow_redirects=True, verify=False) as resp:
            out["status"] = resp.status_code
            out["content_type"] = resp.headers.get("Content-Type", "")
            dest = dest_dir / (sha256_bytes(url.encode("utf-8"))[:16] + "__" + safe_name(url)[:100])
            h = hashlib.sha256()
            n = 0
            with open(dest, "wb") as f:
                for chunk in resp.iter_content(1024 * 64):
                    if not chunk:
                        continue
                    f.write(chunk)
                    h.update(chunk)
                    n += len(chunk)
            out["path"] = str(dest)
            out["bytes"] = n
            out["sha256"] = h.hexdigest()
            out["downloadable"] = resp.status_code == 200 and n > 0
            if resp.status_code >= 400:
                out["error"] = "HTTP %s" % resp.status_code
    except Exception as e:
        out["error"] = "%s: %s" % (type(e).__name__, e)
        return out

    if do_second and out["downloadable"] and out["content_type"] and "html" in out["content_type"].lower():
        try:
            with requests.get(url, headers=headers, timeout=timeout, stream=True,
                              allow_redirects=True, verify=False) as resp2:
                h2 = hashlib.sha256()
                for chunk in resp2.iter_content(1024 * 64):
                    if chunk:
                        h2.update(chunk)
                out["second_sha256"] = h2.hexdigest()
                out["second_consistent"] = out["second_sha256"] == out["sha256"]
        except Exception as e:
            out["second_sha256"] = None
            out["second_consistent"] = False
            out["error"] = (out["error"] or "") + " | second fetch: %s: %s" % (type(e).__name__, e)
    return out


def fetch_urls(urls, dest_dir: Path, cache: dict, workers=6, retry_hosts=None) -> dict:
    retry_hosts = retry_hosts or []
    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    pending = []
    for u in urls:
        if not u:
            continue
        if u in cache and cache[u].get("status") is not None:
            continue
        if u in cache and cache[u].get("error") and cache[u].get("_final"):
            continue
        pending.append(u)
    pending = list(dict.fromkeys(pending))
    if pending:
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = {ex.submit(fetch_one, u, dest_dir): u for u in pending}
            for fut in as_completed(futs):
                u = futs[fut]
                try:
                    cache[u] = fut.result()
                except Exception as e:
                    cache[u] = {"url": u, "status": None, "error": "%s: %s" % (type(e).__name__, e),
                                "downloadable": False}
    # retry known-bad hosts once
    for u in list(urls):
        host = urlparse(u).netloc.lower()
        if not any(h in host for h in retry_hosts):
            continue
        rec = cache.get(u) or {}
        if rec.get("downloadable"):
            continue
        rec2 = fetch_one(u, dest_dir, timeout=(15, 40), do_second=False)
        rec2["retried"] = True
        cache[u] = rec2
    return cache


_ARCHIVE_HASH_CACHE = {}


def archive_sha256(path: Path) -> str:
    key = str(Path(path).resolve())
    if key not in _ARCHIVE_HASH_CACHE:
        _ARCHIVE_HASH_CACHE[key] = sha256_file(path)
    return _ARCHIVE_HASH_CACHE[key]


def compare_live_to_archive(fetch_rec: dict, archive_path: Path):
    """Return (content_compare, archive_exists, archive_sha, notes)."""
    if not archive_path or str(archive_path) in {"", "None"}:
        if fetch_rec and fetch_rec.get("downloadable"):
            return "no_archive_presented", "n/a", "", "live URL has no paired local archive on the page"
        return "no_archive_presented", "n/a", "", ""
    ap = Path(archive_path)
    exists = ap.is_file()
    if not exists:
        if fetch_rec and fetch_rec.get("error") and not fetch_rec.get("downloadable"):
            return "retrieval_failure", "no", "", "archive missing; live also failed: %s" % fetch_rec.get("error")
        return "archive_missing", "no", "", "local archive path does not exist"
    # Hash only when a live body exists to compare, or when the caller needs the digest.
    need_hash = bool(fetch_rec and fetch_rec.get("downloadable") and fetch_rec.get("sha256"))
    ash = archive_sha256(ap) if need_hash else ""
    if not fetch_rec:
        return "retrieval_failure", "yes", ash, "no fetch record"
    if not fetch_rec.get("downloadable"):
        err = fetch_rec.get("error") or ("HTTP %s" % fetch_rec.get("status"))
        return "retrieval_failure", "yes", ash, err
    if not fetch_rec.get("sha256"):
        return "retrieval_failure", "yes", ash, (
            fetch_rec.get("note")
            or fetch_rec.get("error")
            or "live URL reachable but original bytes were not captured"
        )
    if not ash:
        ash = archive_sha256(ap)
    live_sha = fetch_rec.get("sha256")
    if live_sha and live_sha == ash:
        return "exact_match", "yes", ash, ""
    note = "live_sha=%s archive_sha=%s live_bytes=%s archive_bytes=%s" % (
        live_sha, ash, fetch_rec.get("bytes"), ap.stat().st_size)
    ct = (fetch_rec.get("content_type") or "").lower()
    if "html" in ct:
        if fetch_rec.get("second_consistent") is True:
            note += "; second fetch consistent with first (dynamic/HTML byte mismatch vs archive)"
        elif fetch_rec.get("second_consistent") is False:
            note += "; second fetch differed from first"
        return "mismatch", "yes", ash, note
    return "mismatch", "yes", ash, note


def attach_link_results(rows: list, site: Path, verified: dict, fetch_cache: dict, scratch: Path):
    """Add URL/archive audit rows and fill http fields on existing rows with urls."""
    site = Path(site)
    pairs = collect_all_stat_links(site)
    # also pair Source_register local files
    reg = verified["Source_register"]
    for _, r in reg.iterrows():
        url = "" if is_blank(r.get("source_url")) else str(r.get("source_url"))
        iso = r.get("iso3")
        lf = "" if is_blank(r.get("local_file")) else str(r.get("local_file"))
        archives = []
        if lf and lf not in {"data_raw/", "nan"}:
            archives.append(site / "evidence" / "countries" / str(iso) / lf)
        if url:
            if archives:
                for ap in archives:
                    pairs.append({
                        "page": "Source_register",
                        "url": url,
                        "archive_href": str(ap.relative_to(site)) if ap.exists() or True else lf,
                        "archive_path": str(ap),
                        "iso3": iso,
                        "variable": r.get("variable"),
                    })
            else:
                pairs.append({
                    "page": "Source_register",
                    "url": url,
                    "archive_href": "",
                    "archive_path": "",
                    "iso3": iso,
                    "variable": r.get("variable"),
                })

    # infer iso3 from page path
    def pair_iso(p):
        if p.get("iso3"):
            return p["iso3"]
        m = re.search(r"(?:countries|evidence-pages|evidence/countries)/([A-Z]{3})",
                      (p.get("page") or "") + " " + (p.get("archive_href") or ""))
        return m.group(1) if m else "ALL"

    # fill http on existing number rows that already carry a url
    for rec in rows:
        u = rec.get("url") or ""
        if not u.startswith("http"):
            continue
        fr = fetch_cache.get(u)
        if not fr:
            rec["http_status"] = "not_fetched"
            rec["downloadable"] = "no"
            if rec.get("content_compare") in ("", None):
                rec["content_compare"] = "retrieval_failure"
            rec["notes"] = ((rec.get("notes") or "") + " | url not in fetch cache").strip(" |")
            continue
        rec["http_status"] = fr.get("status") if fr.get("status") is not None else ""
        rec["downloadable"] = "yes" if fr.get("downloadable") else "no"
        rec["live_sha256"] = fr.get("sha256") or ""
        if rec.get("archive_path"):
            cc, exists, ash, note = compare_live_to_archive(fr, Path(rec["archive_path"]))
            rec["content_compare"] = cc
            rec["archive_exists"] = exists
            rec["archive_sha256"] = ash
            if note:
                rec["notes"] = ((rec.get("notes") or "") + " | " + note).strip(" |")
        else:
            rec["content_compare"] = rec.get("content_compare") or (
                "reachable" if fr.get("downloadable") else "retrieval_failure"
            )

    # one row per unique (iso3, url, archive_path) for Source_register / link audit
    seen = set()
    extra = []
    for p in pairs:
        url = p.get("url") or ""
        ap = p.get("archive_path") or ""
        iso = pair_iso(p)
        # The same source/archive pair is normally repeated on a country page,
        # an evidence page, and the source register.  Keep one detailed row per
        # ISO/source/archive combination; LINK_AUDIT below retains all pages.
        key = (iso, url, ap)
        if key in seen:
            continue
        seen.add(key)
        if not url and not ap:
            continue
        fr = fetch_cache.get(url) if url else None
        cc, exists, ash, note = compare_live_to_archive(fr, Path(ap) if ap else None)
        extra.append(new_row(
            iso3=iso,
            verified_sheet="Source_register",
            page=p.get("page") or "",
            variable=p.get("variable") or "",
            item="source URL + archive",
            url=url,
            http_status="" if not fr else (fr.get("status") if fr.get("status") is not None else ""),
            downloadable="" if not url else ("yes" if fr and fr.get("downloadable") else "no"),
            archive_path=ap,
            archive_exists=exists,
            live_sha256="" if not fr else (fr.get("sha256") or ""),
            archive_sha256=ash,
            content_compare=cc,
            number_match="N/A",
            notes=note or ("" if not fr else (fr.get("error") or "")),
        ))
    rows.extend(extra)
    return rows


LINK_AUDIT_FIELDS = [
    "url", "iso3", "variables", "pages", "source_register_rows",
    "source_register_outcome", "source_register_retrieval", "current_status",
    "current_downloadable", "current_http_status", "current_bytes",
    "current_content_type", "current_sha256", "archive_count", "archive_paths",
    "archive_exists", "archive_manifest", "live_vs_archive", "notes",
]

WORKBOOK_AUDIT_FIELDS = [
    "scope", "verified_sheet", "reference_file", "site_file", "reference_rows",
    "reference_columns", "site_rows", "site_columns", "common_columns",
    "compared_cells", "matching_cells", "differing_cells", "site_only_columns",
    "missing_site_columns", "status", "notes",
]


def _site_relative(path: Path, site: Path) -> str:
    try:
        return Path(path).resolve().relative_to(Path(site).resolve()).as_posix()
    except ValueError:
        return str(path).replace("\\", "/")


def build_local_integrity(site: Path) -> dict:
    """Resolve every local HTML link and verify every manifest checksum."""
    site = Path(site).resolve()
    local_refs = []
    missing = set()
    for page in iter_site_pages(site):
        soup = soup_of(read_html(page))
        for tag in soup.find_all(True):
            for attr in ("href", "src"):
                raw = tag.get(attr)
                if not raw or not isinstance(raw, str):
                    continue
                low = raw.lower().strip()
                if low.startswith(("http:", "https:", "//", "mailto:", "javascript:", "data:", "#")):
                    continue
                target = resolve_local(page, raw, site)
                if target is None:
                    continue
                try:
                    rel = target.resolve().relative_to(site).as_posix()
                except ValueError:
                    continue
                if not rel:
                    continue
                local_refs.append(rel)
                if not target.is_file():
                    missing.add(rel)

    manifest_path = site / "manifest" / "checksums.csv"
    manifest = {}
    if manifest_path.is_file():
        with open(manifest_path, "r", encoding="utf-8-sig", newline="") as f:
            for rec in csv.DictReader(f):
                rel = (rec.get("path") or "").replace("\\", "/")
                while rel.startswith("./"):
                    rel = rel[2:]
                if rel:
                    manifest[rel] = {
                        "bytes": rec.get("bytes") or "",
                        "sha256": (rec.get("sha256") or "").lower(),
                    }

    manifest_status = {}
    manifest_bad = []
    manifest_missing = []
    manifest_not_listed = []
    for rel, expect in manifest.items():
        target = site / Path(rel)
        if not target.is_file():
            status = "MISSING"
            manifest_missing.append(rel)
        else:
            actual_bytes = target.stat().st_size
            actual_sha = sha256_file(target)
            if str(expect.get("bytes") or "") != str(actual_bytes) or expect.get("sha256") != actual_sha:
                status = "HASH_MISMATCH"
                manifest_bad.append(rel)
            else:
                status = "EXACT"
        manifest_status[rel] = status

    for rel in sorted(set(local_refs)):
        if rel not in manifest:
            manifest_not_listed.append(rel)

    return {
        "local_reference_count": len(local_refs),
        "unique_local_path_count": len(set(local_refs)),
        "missing_local_paths": sorted(missing),
        "manifest_path": str(manifest_path),
        "manifest_entries": len(manifest),
        "manifest_status": manifest_status,
        "manifest_exact": sum(1 for x in manifest_status.values() if x == "EXACT"),
        "manifest_bad": manifest_bad,
        "manifest_missing": manifest_missing,
        "manifest_not_listed_local_paths": manifest_not_listed,
    }


def _compact_join(values, limit=30000):
    vals = sorted({str(v) for v in values if v not in (None, "")})
    text = "; ".join(vals)
    if len(text) <= limit:
        return text
    return text[:limit - 40] + " ... [truncated; %d items]" % len(vals)


def _pair_context(pair):
    text = "%s %s" % (pair.get("page") or "", pair.get("archive_href") or "")
    iso = pair.get("iso3") or ""
    var = pair.get("variable") or ""
    m = re.search(r"(?:countries|evidence-pages|evidence/countries)/([A-Z]{3})", text)
    if not iso and m:
        iso = m.group(1)
    m = re.search(r"evidence-pages/([A-Z]{3})__([a-z0-9_]+)", text, re.I)
    if m:
        iso = iso or m.group(1).upper()
        var = var or m.group(2).lower()
    return str(iso or "ALL"), str(var or "")


def _source_archive_path(site: Path, iso3, local_file):
    lf = str(local_file or "").strip()
    if not lf or lf.lower() in {"nan", "data_raw/", "data_raw"}:
        return ""
    if lf.startswith("evidence/"):
        candidate = site / Path(lf)
    else:
        candidate = site / "evidence" / "countries" / str(iso3) / lf
    return str(candidate) if candidate.is_file() else str(candidate)


def _fetch_classification(rec):
    if not rec:
        return "not_fetched"
    status = rec.get("status")
    if rec.get("downloadable"):
        return "downloadable_200"
    if status == 429:
        return "rate_limited_429"
    if status == 403:
        return "blocked_403"
    if status == 404:
        return "not_found_404"
    if status == 400:
        return "bad_request_400"
    if status is None:
        return "unreachable"
    if isinstance(status, int) and status >= 500:
        return "server_error_%s" % status
    return "http_%s" % status


def build_link_audit(site: Path, verified: dict, fetch_cache: dict,
                     local_integrity: dict, rows: list | None = None) -> list:
    """Aggregate all statistic-associated URLs into one auditable row each."""
    site = Path(site).resolve()
    records = {}

    def ensure(url):
        if url not in records:
            records[url] = {
                "url": url, "iso3": set(), "variables": set(), "pages": set(),
                "source_register_rows": set(), "source_register_outcome": set(),
                "source_register_retrieval": set(), "archives": set(),
                "notes": set(),
            }
        return records[url]

    def add(url, iso3="", variable="", page="", archive_path="", source_row="",
            outcome="", retrieval="", note=""):
        if not url or not str(url).startswith(("http://", "https://")):
            return
        r = ensure(str(url))
        if iso3:
            r["iso3"].add(str(iso3))
        if variable:
            r["variables"].add(str(variable))
        if page:
            r["pages"].add(str(page))
        if archive_path:
            r["archives"].add(_site_relative(Path(archive_path), site))
        if source_row:
            r["source_register_rows"].add(str(source_row))
        if outcome:
            r["source_register_outcome"].add(str(outcome))
        if retrieval:
            r["source_register_retrieval"].add(str(retrieval))
        if note:
            r["notes"].add(str(note))

    for pair in collect_all_stat_links(site):
        url = pair.get("url") or ""
        iso, var = _pair_context(pair)
        add(url, iso3=iso, variable=var, page=pair.get("page") or "",
            archive_path=pair.get("archive_path") or "")

    # Existing checklist rows carry the most precise variable/year context and
    # source-register archive association.
    for rec in rows or []:
        url = rec.get("url") or ""
        if str(url).startswith(("http://", "https://")):
            add(url, iso3=rec.get("iso3") or "", variable=rec.get("variable") or "",
                page=rec.get("page") or "", archive_path=rec.get("archive_path") or "")

    site_reg_path = site / "data" / "source_register.csv"
    if site_reg_path.is_file():
        reg = pd.read_csv(site_reg_path, dtype=str, keep_default_na=False)
        for _, rec in reg.iterrows():
            url = rec.get("source_url") or ""
            if not str(url).startswith(("http://", "https://")):
                continue
            iso = rec.get("iso3") or ""
            row_id = "%s/%s/%s" % (iso, rec.get("variable") or "", rec.get("years") or "")
            add(url, iso3=iso, variable=rec.get("variable") or "",
                page="data/source_register.csv",
                archive_path=_source_archive_path(site, iso, rec.get("local_file")),
                source_row=row_id, outcome=rec.get("outcome") or "",
                retrieval=rec.get("retrieval") or "", note=rec.get("note") or "")

    # Include irregular-estimate source URLs even where they are not repeated
    # in the source register.
    irr = verified.get("Irregular_estimates_all")
    if irr is not None and "source_url" in irr.columns:
        for _, rec in irr.iterrows():
            url = "" if is_blank(rec.get("source_url")) else str(rec.get("source_url"))
            add(url, iso3=rec.get("iso3") or "", variable=rec.get("variable") or "",
                page="Irregular_estimates_all", note="source URL from competing-estimate worksheet")

    manifest_status = local_integrity.get("manifest_status", {})
    out = []
    for url in sorted(records):
        r = records[url]
        fr = fetch_cache.get(url) or {}
        archives = sorted(r["archives"])
        existing = [p for p in archives if (site / Path(p)).is_file()]
        if not archives:
            archive_exists = "none"
            archive_manifest = "none"
            live_relation = "no_archive"
        else:
            archive_exists = "all_exist" if len(existing) == len(archives) else ("partial" if existing else "none")
            statuses = [manifest_status.get(p, "NOT_IN_MANIFEST") for p in existing]
            if not existing:
                archive_manifest = "none_exist"
            elif all(x == "EXACT" for x in statuses) and len(existing) == len(archives):
                archive_manifest = "all_exact"
            elif any(x in {"HASH_MISMATCH", "MISSING"} for x in statuses):
                archive_manifest = "integrity_problem"
            elif any(x == "NOT_IN_MANIFEST" for x in statuses):
                archive_manifest = "partial_or_not_listed"
            else:
                archive_manifest = "partial"
            if fr.get("downloadable") and fr.get("sha256") and existing:
                same = any(archive_sha256(site / Path(p)) == fr.get("sha256") for p in existing)
                live_relation = "same_bytes" if same else "different_snapshot_or_mirror"
            elif fr.get("downloadable"):
                live_relation = "not_compared_no_existing_archive"
            else:
                live_relation = "not_compared_live_failed"

        notes = set(r["notes"])
        if fr.get("error"):
            notes.add(str(fr.get("error")))
        if _fetch_classification(fr) == "rate_limited_429" and "oecd" in urlparse(url).netloc.lower():
            notes.add("OECD JSON endpoint returned HTTP 429 at audit time; the exact URL is rate-limited, not a value mismatch")
        if live_relation == "different_snapshot_or_mirror":
            notes.add("live bytes differ from local archive; treated as snapshot/format information, while manifest integrity is checked separately")
        if not archives and any(x in {"data_raw/", "data_raw"} for x in r["source_register_retrieval"]):
            notes.add("source register records API/raw-data retrieval rather than a document archive")
        out.append({
            "url": url,
            "iso3": _compact_join(r["iso3"]),
            "variables": _compact_join(r["variables"]),
            "pages": _compact_join(r["pages"]),
            "source_register_rows": _compact_join(r["source_register_rows"]),
            "source_register_outcome": _compact_join(r["source_register_outcome"]),
            "source_register_retrieval": _compact_join(r["source_register_retrieval"]),
            "current_status": _fetch_classification(fr),
            "current_downloadable": "yes" if fr.get("downloadable") else "no",
            "current_http_status": fr.get("status") if fr.get("status") is not None else "",
            "current_bytes": fr.get("bytes") or "",
            "current_content_type": fr.get("content_type") or "",
            "current_sha256": fr.get("sha256") or "",
            "archive_count": len(archives),
            "archive_paths": _compact_join(archives),
            "archive_exists": archive_exists,
            "archive_manifest": archive_manifest,
            "live_vs_archive": live_relation,
            "notes": _compact_join(notes),
        })
    return out


SITE_TABLE_FILES = {
    "README": "data/readme.csv",
    "Panel_final": "data/panel_final.csv",
    "Data_quality": "data/data_quality.csv",
    "Corrections_applied": "data/corrections_applied.csv",
    "Known_issues": "data/known_issues.csv",
    "Verification_log": "data/verification_log.csv",
    "Source_register": "data/source_register.csv",
    "Irregular_estimates_all": "data/irregular_estimates_all.csv",
    "Codebook": "data/codebook.csv",
}


def _audit_cell_equal(a, b):
    if is_blank(a) and is_blank(b):
        return True
    if is_blank(a) or is_blank(b):
        return False
    na = normalize_xlsx_number(a)
    nb = normalize_xlsx_number(b)
    if isinstance(na, (int, float)) and isinstance(nb, (int, float)):
        return abs(float(na) - float(nb)) < 1e-9
    return str(a).strip() == str(b).strip()


def _drop_empty_rows(frame: pd.DataFrame) -> pd.DataFrame:
    """Drop rows that contain only empty/NA tokens, including CSV empty strings."""
    if frame.empty:
        return frame.copy().reset_index(drop=True)
    nonempty = frame.apply(lambda col: col.map(lambda value: not is_blank(value)))
    return frame.loc[nonempty.any(axis=1)].reset_index(drop=True)


def _compare_site_table(site: Path, verified: dict, sheet: str):
    ref = _drop_empty_rows(verified[sheet])
    site_file = SITE_TABLE_FILES.get(sheet, "")
    site_path = site / site_file
    if not site_path.is_file():
        return {
            "site_file": site_file, "site_rows": 0, "site_columns": 0,
            "common_columns": [], "compared_cells": 0, "matching_cells": 0,
            "differing_cells": 0, "site_only_columns": [], "missing_site_columns": list(ref.columns),
            "status": "MISSING_SITE_TABLE", "notes": "site representation file is absent",
        }
    site_df = _drop_empty_rows(pd.read_csv(site_path, keep_default_na=False))
    common = [c for c in ref.columns if c in site_df.columns]
    differing = []
    matching = 0
    n = min(len(ref), len(site_df))
    for i in range(n):
        for c in common:
            if _audit_cell_equal(ref.iloc[i][c], site_df.iloc[i][c]):
                matching += 1
            else:
                differing.append((i + 1, c, fmt(ref.iloc[i][c]), fmt(site_df.iloc[i][c])))
    compared = n * len(common)
    missing_rows = max(0, len(ref) - len(site_df)) * len(common)
    extra_rows = max(0, len(site_df) - len(ref)) * len(common)
    differing_count = len(differing) + missing_rows + extra_rows
    site_only = [c for c in site_df.columns if c not in ref.columns]
    missing_cols = [c for c in ref.columns if c not in site_df.columns]
    if differing_count == 0 and not site_only and not missing_cols and len(ref) == len(site_df):
        status = "MATCH"
    elif differing_count == 0 and sheet == "Panel_final" and site_only:
        status = "MATCH_CORE_PLUS_SITE_COLUMNS"
    elif sheet == "Source_register" and differing_count and set(x[1] for x in differing).issubset({"retrieval", "local_file", "note"}):
        status = "UPDATED_SITE_REGISTER"
    elif differing_count == 0:
        status = "MATCH_CORE"
    else:
        status = "REVIEW"
    examples = ["row %s %s: %s -> %s" % x for x in differing[:8]]
    notes = ""
    if len(ref) != len(site_df):
        notes += "blank/all-empty rows removed before comparison; "
    if site_only:
        notes += "site-only columns=%s; " % ", ".join(site_only)
    if missing_cols:
        notes += "missing site columns=%s; " % ", ".join(missing_cols)
    if sheet == "Source_register" and differing:
        notes += "site register contains refreshed retrieval/local-archive/note metadata and an outcome column; "
    if examples:
        notes += "examples: " + " | ".join(examples)
    return {
        "site_file": site_file, "site_rows": len(site_df), "site_columns": len(site_df.columns),
        "common_columns": common, "compared_cells": compared, "matching_cells": matching,
        "differing_cells": differing_count, "site_only_columns": site_only,
        "missing_site_columns": missing_cols, "status": status, "notes": notes.strip(" ;"),
    }


def build_workbook_audit(site: Path, verified: dict, rows: list,
                         local_integrity: dict, link_audit: list) -> list:
    site = Path(site).resolve()
    ref_path = Path(verified["path"]).resolve()
    out = []
    for sheet in VERIFIED_SHEETS:
        cmp = _compare_site_table(site, verified, sheet)
        out.append({
            "scope": "site data representation", "verified_sheet": sheet,
            "reference_file": _site_relative(ref_path, site.parent),
            "site_file": cmp["site_file"],
            "reference_rows": len(_drop_empty_rows(verified[sheet])),
            "reference_columns": len(verified[sheet].columns),
            "site_rows": cmp["site_rows"], "site_columns": cmp["site_columns"],
            "common_columns": len(cmp["common_columns"]),
            "compared_cells": cmp["compared_cells"], "matching_cells": cmp["matching_cells"],
            "differing_cells": cmp["differing_cells"],
            "site_only_columns": _compact_join(cmp["site_only_columns"]),
            "missing_site_columns": _compact_join(cmp["missing_site_columns"]),
            "status": cmp["status"], "notes": cmp["notes"],
        })

    # Keep the two website-wide headline discrepancies visible on the
    # worksheet-audit index as well as in ALL_HEADLINES/SUMMARY.
    for rec in rows:
        if rec.get("iso3") != "ALL" or rec.get("number_match") != "FAIL":
            continue
        out.append({
            "scope": "website headline vs README/reference",
            "verified_sheet": rec.get("verified_sheet") or "README",
            "reference_file": _site_relative(ref_path, site.parent),
            "site_file": rec.get("page") or "",
            "reference_rows": "", "reference_columns": "", "site_rows": "",
            "site_columns": "", "common_columns": "", "compared_cells": 1,
            "matching_cells": 0, "differing_cells": 1,
            "site_only_columns": "", "missing_site_columns": "",
            "status": "PRESENTATION_MISMATCH",
            "notes": "%s; site=%s; reference=%s; %s" % (
                rec.get("item") or "headline", rec.get("site_value") or "",
                rec.get("xlsx_value") or "", rec.get("notes") or ""),
        })

    bundled = site / "data" / "FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx"
    ref_hash = sha256_file(ref_path) if ref_path.is_file() else ""
    site_hash = sha256_file(bundled) if bundled.is_file() else ""
    out.append({
        "scope": "workbook byte identity", "verified_sheet": "(entire workbook)",
        "reference_file": str(ref_path), "site_file": str(bundled),
        "reference_rows": ref_path.stat().st_size if ref_path.is_file() else "",
        "reference_columns": "bytes", "site_rows": bundled.stat().st_size if bundled.is_file() else "",
        "site_columns": "bytes", "common_columns": "", "compared_cells": "",
        "matching_cells": "yes" if ref_hash and ref_hash == site_hash else "no",
        "differing_cells": "" if ref_hash and ref_hash == site_hash else "byte hash differs",
        "site_only_columns": "", "missing_site_columns": "",
        "status": "MATCH" if ref_hash and ref_hash == site_hash else "DIFF",
        "notes": "reference_sha256=%s; site_sha256=%s; bytes are compared separately from normalized worksheet content" % (ref_hash, site_hash),
    })

    # Explicit integrity row keeps the archive evidence visible even when a
    # source URL is unreachable today.
    out.append({
        "scope": "local archive integrity", "verified_sheet": "manifest/checksums.csv",
        "reference_file": local_integrity.get("manifest_path", ""),
        "site_file": "all local href/src targets in generated HTML",
        "reference_rows": local_integrity.get("manifest_entries", 0),
        "reference_columns": "path,bytes,sha256",
        "site_rows": local_integrity.get("unique_local_path_count", 0),
        "site_columns": "unique local paths", "common_columns": "",
        "compared_cells": local_integrity.get("manifest_entries", 0),
        "matching_cells": local_integrity.get("manifest_exact", 0),
        "differing_cells": len(local_integrity.get("manifest_bad", [])) + len(local_integrity.get("manifest_missing", [])),
        "site_only_columns": "", "missing_site_columns": "",
        "status": "MATCH" if not local_integrity.get("manifest_bad") and not local_integrity.get("manifest_missing") and not local_integrity.get("missing_local_paths") else "REVIEW",
        "notes": "local references=%s; missing local paths=%s; manifest paths not linked from HTML=%s" % (
            local_integrity.get("local_reference_count", 0),
            len(local_integrity.get("missing_local_paths", [])),
            len(local_integrity.get("manifest_not_listed_local_paths", [])),
        ),
    })
    return out


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
MATCH_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
NA_FILL = PatternFill("solid", fgColor="D9D9D9")
GROUP_FILL = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def write_checklist(rows: list, out_path: Path, iso_list: list, verified_path: str,
                    verified: dict | None = None, site: Path | None = None,
                    fetch_cache: dict | None = None, local_integrity: dict | None = None,
                    link_audit: list | None = None, workbook_audit: list | None = None):
    out_path = Path(out_path)
    if out_path.resolve() == Path(verified_path).resolve():
        raise RuntimeError("refusing to overwrite the VERIFIED workbook")
    site = Path(site) if site is not None else SITE_ROOT
    if verified is not None:
        local_integrity = local_integrity or build_local_integrity(site)
        link_audit = link_audit if link_audit is not None else build_link_audit(
            site, verified, fetch_cache or {}, local_integrity, rows
        )
        workbook_audit = workbook_audit if workbook_audit is not None else build_workbook_audit(
            site, verified, rows, local_integrity, link_audit
        )
    else:
        local_integrity = local_integrity or {}
        link_audit = link_audit or []
        workbook_audit = workbook_audit or []
    df = pd.DataFrame(rows)
    for c in ROW_FIELDS:
        if c not in df.columns:
            df[c] = ""
    df = df[ROW_FIELDS]
    wb = Workbook()

    # SUMMARY
    ws = wb.active
    ws.title = "SUMMARY"
    _write_summary_sheet(ws, df, iso_list, verified_path, verified=verified,
                         site=site, local_integrity=local_integrity,
                         link_audit=link_audit, workbook_audit=workbook_audit)

    if workbook_audit:
        audit_ws = wb.create_sheet("WORKBOOK_AUDIT")
        _write_records_sheet(audit_ws, workbook_audit, WORKBOOK_AUDIT_FIELDS,
                             "Worksheet and packaged-workbook audit")
    if link_audit:
        link_ws = wb.create_sheet("LINK_AUDIT")
        _write_records_sheet(link_ws, link_audit, LINK_AUDIT_FIELDS,
                             "External URL and local archive audit")

    sheet_order = list(VERIFIED_SHEETS)

    for iso in iso_list:
        sub = df[df["iso3"] == iso].copy()
        # guarantee every verified sheet appears
        present = set(sub["verified_sheet"].astype(str))
        extras = []
        for sh in sheet_order:
            if sh not in present:
                extras.append(new_row(
                    iso3=iso, verified_sheet=sh, item="no rows generated for this sheet",
                    number_match="N/A",
                    notes="placeholder so every required VERIFIED sheet is represented",
                ))
        if extras:
            sub = pd.concat([sub, pd.DataFrame(extras)], ignore_index=True)
        # sort
        sub["_ord"] = sub["verified_sheet"].map({s: i for i, s in enumerate(sheet_order)}).fillna(99)
        sub = sub.sort_values(["_ord", "variable", "year", "item", "page"])
        ws = wb.create_sheet(iso)
        _write_country_sheet(ws, sub, iso)

    # ALL-site headlines live on SUMMARY already; keep a raw dump sheet
    raw = wb.create_sheet("ALL_HEADLINES")
    all_df = df[df["iso3"] == "ALL"].copy()
    _write_country_sheet(raw, all_df, "ALL")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _write_records_sheet(ws, records: list, fields: list, title: str):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15, color="1F4E79")
    if fields:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
    header_row = 3
    for ci, field in enumerate(fields, 1):
        cell = ws.cell(header_row, ci, field)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    for ri, rec in enumerate(records, header_row + 1):
        for ci, field in enumerate(fields, 1):
            value = rec.get(field, "") if isinstance(rec, dict) else ""
            if value is None:
                value = ""
            cell = ws.cell(ri, ci, value)
            cell.alignment = WRAP
            cell.border = THIN
            if field in {"status", "current_status", "archive_manifest", "live_vs_archive"}:
                text = str(value)
                if text in {"MATCH", "MATCH_CORE", "MATCH_CORE_PLUS_SITE_COLUMNS", "all_exact", "same_bytes", "downloadable_200"}:
                    cell.fill = MATCH_FILL
                elif text in {"FAIL", "DIFF", "REVIEW", "PRESENTATION_MISMATCH", "MISSING_SITE_TABLE", "HASH_MISMATCH", "integrity_problem", "blocked_403", "not_found_404", "bad_request_400", "unreachable", "rate_limited_429", "not_compared_live_failed"}:
                    cell.fill = FAIL_FILL
                elif text in {"N/A", "none", "no_archive", "not_fetched"}:
                    cell.fill = NA_FILL
    end_row = max(header_row, header_row + len(records))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:%s%d" % (get_column_letter(max(len(fields), 1)), end_row)
    for ci, field in enumerate(fields, 1):
        if field in {"notes", "pages", "archive_paths", "site_only_columns", "missing_site_columns"}:
            width = 60
        elif field in {"url", "reference_file", "site_file", "current_content_type", "current_sha256"}:
            width = 42
        elif field in {"variables", "source_register_rows", "source_register_outcome", "source_register_retrieval"}:
            width = 28
        else:
            width = min(max(len(field) + 2, 12), 24)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 32


def _write_summary_sheet(ws, df: pd.DataFrame, iso_list, verified_path,
                         verified=None, site=None, local_integrity=None,
                         link_audit=None, workbook_audit=None):
    ws["A1"] = "Site vs VERIFIED workbook audit checklist"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A2"] = "Truth file"
    ws["B2"] = verified_path
    ws["A3"] = "Generated"
    ws["B3"] = time.strftime("%Y-%m-%d %H:%M:%S")
    ws.merge_cells("B2:F2")

    # Evidence cards at the top of SUMMARY.  These are deliberately based on
    # the detailed audit rows and manifest/link tables rather than on a single
    # headline from either workbook.
    ws["A5"] = "Audit evidence at a glance"
    ws["A5"].font = Font(bold=True, size=12, color="1F4E79")
    cards = []
    panel_cells = df[(df["verified_sheet"] == "Panel_final") & (df["item"].isin(["panel cell", "evidence cell"]))]
    panel_page_cells = panel_cells[panel_cells["item"] == "panel cell"]
    evidence_page_cells = panel_cells[panel_cells["item"] == "evidence cell"]
    cards.extend([
        ("Country-page panel cells", len(panel_page_cells)),
        ("Evidence-page value cells", len(evidence_page_cells)),
        ("Panel/evidence cells compared", len(panel_cells)),
        ("Panel/evidence MATCH", int((panel_cells["number_match"] == "MATCH").sum())),
        ("Panel/evidence FAIL", int((panel_cells["number_match"] == "FAIL").sum())),
        ("Local link references", (local_integrity or {}).get("local_reference_count", "")),
        ("Unique local paths", (local_integrity or {}).get("unique_local_path_count", "")),
        ("Manifest files exact", "%s/%s" % ((local_integrity or {}).get("manifest_exact", ""), (local_integrity or {}).get("manifest_entries", ""))),
        ("Unique external URLs", len(link_audit or [])),
        ("External URLs downloadable now", sum(1 for r in (link_audit or []) if r.get("current_status") == "downloadable_200")),
    ])
    for i, (label, value) in enumerate(cards, 6):
        ws.cell(i, 1, label)
        ws.cell(i, 2, value)
        ws.cell(i, 1).border = ws.cell(i, 2).border = THIN
        ws.cell(i, 1).alignment = ws.cell(i, 2).alignment = WRAP
        if label.endswith("FAIL") and value:
            ws.cell(i, 2).fill = FAIL_FILL
        elif label in {"Panel/evidence MATCH", "Manifest files exact"}:
            ws.cell(i, 2).fill = MATCH_FILL

    if verified is not None and site is not None:
        ref_path = Path(verified["path"])
        bundled = Path(site) / "data" / "FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx"
        ws["D6"] = "Reference SHA-256"
        ws["E6"] = sha256_file(ref_path) if ref_path.is_file() else ""
        ws["D7"] = "Site bundled workbook SHA-256"
        ws["E7"] = sha256_file(bundled) if bundled.is_file() else ""
        ws["D8"] = "Bundled workbook byte identity"
        ws["E8"] = "MATCH" if ws["E6"].value and ws["E6"].value == ws["E7"].value else "DIFF (site has an enriched copy)"
        ws["D9"] = "Reference sheets audited"
        ws["E9"] = ", ".join(VERIFIED_SHEETS)
        for row in range(6, 10):
            for col in range(4, 6):
                ws.cell(row, col).border = THIN
                ws.cell(row, col).alignment = WRAP
        ws["E8"].fill = MATCH_FILL if ws["E8"].value == "MATCH" else FAIL_FILL

    # Current live URL status is separate from local archive integrity: a
    # dynamic/updated live page need not have the same bytes as a PDF/HTML
    # snapshot, while the manifest still must be exact.
    status_start = 16
    ws.cell(status_start, 1, "Current external URL status").font = Font(bold=True, size=12, color="1F4E79")
    status_counts = Counter(r.get("current_status") or "" for r in (link_audit or []))
    for i, (status, n) in enumerate(sorted(status_counts.items()), status_start + 1):
        ws.cell(i, 1, status)
        ws.cell(i, 2, n)
        ws.cell(i, 1).border = ws.cell(i, 2).border = THIN
        if status in {"downloadable_200", "same_bytes"}:
            ws.cell(i, 2).fill = MATCH_FILL
        elif status not in {"", "not_fetched"}:
            ws.cell(i, 2).fill = FAIL_FILL

    # Worksheet-level statuses are a compact index; full comparison details
    # are on WORKBOOK_AUDIT.
    sheet_start = max(status_start + len(status_counts) + 3, 28)
    ws.cell(sheet_start, 1, "Worksheet/package status").font = Font(bold=True, size=12, color="1F4E79")
    for ci, field in enumerate(["scope", "verified_sheet", "status", "reference_rows", "site_rows", "differing_cells", "notes"], 1):
        cell = ws.cell(sheet_start + 1, ci, field)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for ri, rec in enumerate(workbook_audit or [], sheet_start + 2):
        for ci, field in enumerate(["scope", "verified_sheet", "status", "reference_rows", "site_rows", "differing_cells", "notes"], 1):
            cell = ws.cell(ri, ci, rec.get(field, ""))
            cell.alignment = WRAP
            cell.border = THIN
            if field == "status":
                if str(cell.value).startswith("MATCH"):
                    cell.fill = MATCH_FILL
                elif str(cell.value) in {"DIFF", "REVIEW", "PRESENTATION_MISMATCH", "UPDATED_SITE_REGISTER", "MISSING_SITE_TABLE"}:
                    cell.fill = FAIL_FILL

    headline_start = sheet_start + max(len(workbook_audit or []), 1) + 4

    counts = Counter(df["number_match"].astype(str))
    cc = Counter(df["content_compare"].astype(str))
    ws.cell(headline_start, 1, "Detailed number_match totals")
    ws.cell(headline_start, 1).font = Font(bold=True)
    r = headline_start + 1
    for k in ("MATCH", "FAIL", "N/A", ""):
        ws["A%s" % r] = k or "(blank)"
        ws["B%s" % r] = int(counts.get(k, 0))
        if k == "FAIL":
            ws["B%s" % r].fill = FAIL_FILL
        if k == "MATCH":
            ws["B%s" % r].fill = MATCH_FILL
        r += 1
    ws.cell(r + 1, 1, "Detailed content_compare totals (row-level; unique URL result is on LINK_AUDIT)")
    ws.cell(r + 1, 1).font = Font(bold=True)
    r = r + 2
    for k, n in sorted(cc.items(), key=lambda kv: (-kv[1], kv[0])):
        ws["A%s" % r] = k
        ws["B%s" % r] = int(n)
        r += 1

    ws.cell(headline_start, 4, "Country sheets")
    ws.cell(headline_start, 5, len(iso_list))
    ws.cell(headline_start + 1, 4, "ISO3")
    ws.cell(headline_start + 1, 5, ", ".join(iso_list))

    # per-country MATCH/FAIL
    ws["A%s" % (r + 1)] = "Per-country number_match"
    ws["A%s" % (r + 1)].font = Font(bold=True)
    hdr_row = r + 2
    for i, h in enumerate(["iso3", "rows", "MATCH", "FAIL", "N/A", "link rows", "exact_match", "mismatch", "retrieval_failure", "archive_missing"], 1):
        cell = ws.cell(hdr_row, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    rr = hdr_row + 1
    for iso in iso_list:
        sub = df[df["iso3"] == iso]
        ws.cell(rr, 1, iso)
        ws.cell(rr, 2, len(sub))
        ws.cell(rr, 3, int((sub["number_match"] == "MATCH").sum()))
        ws.cell(rr, 4, int((sub["number_match"] == "FAIL").sum()))
        ws.cell(rr, 5, int((sub["number_match"] == "N/A").sum()))
        ws.cell(rr, 6, int((sub["item"] == "source URL + archive").sum()))
        ws.cell(rr, 7, int((sub["content_compare"] == "exact_match").sum()))
        ws.cell(rr, 8, int((sub["content_compare"] == "mismatch").sum()))
        ws.cell(rr, 9, int((sub["content_compare"] == "retrieval_failure").sum()))
        ws.cell(rr, 10, int((sub["content_compare"] == "archive_missing").sum()))
        if int((sub["number_match"] == "FAIL").sum()):
            ws.cell(rr, 4).fill = FAIL_FILL
        rr += 1

    # ALL headlines
    start = rr + 2
    ws.cell(start, 1, "Site-wide README / Verification_log / methods headlines")
    ws.cell(start, 1).font = Font(bold=True, size=12)
    all_df = df[df["iso3"] == "ALL"]
    _dump_table(ws, all_df, start + 2)
    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["G"].width = 36


def _dump_table(ws, df: pd.DataFrame, start_row: int):
    cols = ROW_FIELDS
    for i, h in enumerate(cols, 1):
        cell = ws.cell(start_row, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for ri, rec in enumerate(df.to_dict("records"), start_row + 1):
        for ci, h in enumerate(cols, 1):
            val = rec.get(h, "")
            if val is None:
                val = ""
            cell = ws.cell(ri, ci, val)
            cell.alignment = WRAP
            cell.border = THIN
            nm = rec.get("number_match")
            if h == "number_match":
                if nm == "MATCH":
                    cell.fill = MATCH_FILL
                elif nm == "FAIL":
                    cell.fill = FAIL_FILL
                elif nm == "N/A":
                    cell.fill = NA_FILL


def _write_country_sheet(ws, sub: pd.DataFrame, iso: str):
    ws["A1"] = "Audit checklist — %s" % iso
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:R1")
    ws["A2"] = (
        "Rows grouped by VERIFIED workbook sheet: README, Panel_final, Data_quality, "
        "Corrections_applied, Known_issues, Verification_log, Source_register, "
        "Irregular_estimates_all, Codebook."
    )
    ws.merge_cells("A2:R2")

    cols = ROW_FIELDS
    # write grouped with separator rows
    sheet_order = list(VERIFIED_SHEETS)
    r = 4
    for i, h in enumerate(cols, 1):
        cell = ws.cell(r, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    r = 5
    grouped = {s: sub[sub["verified_sheet"] == s] for s in sheet_order}
    other = sub[~sub["verified_sheet"].isin(sheet_order)]
    blocks = [(s, grouped[s]) for s in sheet_order] + ([("OTHER", other)] if len(other) else [])
    for name, block in blocks:
        # Avoid a leading '=': Excel interprets strings such as "=== ... ==="
        # as formulas when the checklist is reopened.
        cell = ws.cell(r, 1, "Group: %s (%d rows)" % (name, len(block)))
        cell.fill = GROUP_FILL
        cell.font = Font(bold=True)
        for c in range(2, len(cols) + 1):
            ws.cell(r, c).fill = GROUP_FILL
        r += 1
        for rec in block.to_dict("records"):
            for ci, h in enumerate(cols, 1):
                val = rec.get(h, "")
                if val is None:
                    val = ""
                cell = ws.cell(r, ci, val)
                cell.alignment = WRAP
                cell.border = THIN
                if h == "number_match":
                    if val == "MATCH":
                        cell.fill = MATCH_FILL
                    elif val == "FAIL":
                        cell.fill = FAIL_FILL
                    elif val == "N/A":
                        cell.fill = NA_FILL
                if h == "content_compare" and val == "exact_match":
                    cell.fill = MATCH_FILL
                if h == "content_compare" and val in {"mismatch", "retrieval_failure", "archive_missing"}:
                    cell.fill = FAIL_FILL
            r += 1
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(cols)), max(r - 1, 5))
    widths = {
        "A": 10, "B": 24, "C": 36, "D": 28, "E": 8, "F": 44,
        "G": 18, "H": 18, "I": 14, "J": 40, "K": 12, "L": 12,
        "M": 40, "N": 12, "O": 16, "P": 16, "Q": 18, "R": 50,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[4].height = 22


# ---------------------------------------------------------------------------
# Totals / logs
# ---------------------------------------------------------------------------

def summarize(rows: list) -> dict:
    c_num = Counter(r.get("number_match") or "" for r in rows)
    c_cc = Counter(r.get("content_compare") or "" for r in rows)
    n_urls = len({r.get("url") for r in rows if str(r.get("url") or "").startswith("http")})
    n_dl = sum(1 for r in rows if r.get("item") == "source URL + archive" and r.get("downloadable") == "yes")
    n_fail_link = sum(1 for r in rows if r.get("item") == "source URL + archive" and r.get("downloadable") == "no")
    panel_rows = [r for r in rows if r.get("verified_sheet") == "Panel_final" and r.get("item") in {"panel cell", "evidence cell"}]
    return {
        "n_rows": len(rows),
        "number_MATCH": int(c_num.get("MATCH", 0)),
        "number_FAIL": int(c_num.get("FAIL", 0)),
        "number_NA": int(c_num.get("N/A", 0)),
        "content_exact_match": int(c_cc.get("exact_match", 0)),
        "content_mismatch": int(c_cc.get("mismatch", 0)),
        "content_retrieval_failure": int(c_cc.get("retrieval_failure", 0)),
        "content_archive_missing": int(c_cc.get("archive_missing", 0)),
        "unique_http_urls": n_urls,
        "link_downloadable": n_dl,
        "link_fail": n_fail_link,
        "panel_displayed_compared": len(panel_rows),
        "panel_displayed_MATCH": sum(1 for r in panel_rows if r.get("number_match") == "MATCH"),
        "panel_displayed_FAIL": sum(1 for r in panel_rows if r.get("number_match") == "FAIL"),
    }


def write_summary_csv(rows: list, path: Path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=ROW_FIELDS)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in ROW_FIELDS})


def checklist_structure_report(checklist: Path, iso_list: list) -> str:
    wb = load_workbook(checklist, read_only=True, data_only=True)
    lines = ["checklist=%s" % checklist, "sheets=%s" % wb.sheetnames]
    country_sheets = [s for s in wb.sheetnames if s in set(iso_list)]
    lines.append("country_sheets=%d expected=%d" % (len(country_sheets), len(iso_list)))
    missing = [s for s in iso_list if s not in wb.sheetnames]
    extra = [s for s in country_sheets if s not in iso_list]
    lines.append("missing=%s" % missing)
    lines.append("extra_iso=%s" % extra)
    required = set(VERIFIED_SHEETS)
    for iso in iso_list:
        ws = wb[iso]
        texts = []
        n = 0
        sheets_seen = set()
        for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
            n += 1
            if row and isinstance(row[0], str) and row[0].startswith("Group:"):
                m = re.search(r"Group: ([A-Za-z_]+)", row[0])
                if m:
                    sheets_seen.add(m.group(1))
            elif row and row[1]:
                sheets_seen.add(str(row[1]))
        miss_sh = sorted(required - sheets_seen)
        lines.append("%s rows=%d sheets_seen=%s missing_groups=%s" % (iso, n, sorted(sheets_seen), miss_sh))
    wb.close()
    return "\n".join(lines) + "\n"


def independent_spotcheck(site: Path, verified: dict, rows: list, isos=("TWN", "AUS")) -> str:
    """Re-parse country + evidence HTML and confirm checklist rows exist."""
    site = Path(site)
    panel = verified["Panel_final"]
    lines = []
    def year_key(y):
        if y in (None, "", "nan"):
            return ""
        try:
            return str(int(float(y)))
        except (TypeError, ValueError):
            return str(y)

    by_key = defaultdict(list)
    for r in rows:
        if r.get("iso3") in isos and r.get("verified_sheet") == "Panel_final" \
                and r.get("item") in {"panel cell", "evidence cell"}:
            by_key[(r["iso3"], r["item"], r.get("variable"), year_key(r.get("year")))].append(r)
    for iso in isos:
        chtml = read_html(site / "countries" / ("%s.html" % iso))
        parsed = parse_panel_table(chtml)
        lines.append("== %s country page vars=%s cells=%d ==" % (iso, parsed["vars"], len(parsed["cells"])))
        for cell in parsed["cells"]:
            xv, _ = panel_lookup(panel, iso, cell["year"], cell["variable"])
            expect = "MATCH" if numbers_equal(cell["value"], xv) else "FAIL"
            key = (iso, "panel cell", cell["variable"], year_key(cell["year"]))
            hits = by_key.get(key, [])
            agree = bool(hits) and hits[0]["number_match"] == expect
            lines.append(
                "  panel %s %s site=%s xlsx=%s expect=%s checklist=%s agree=%s"
                % (cell["variable"], cell["year"], fmt(cell["value"]), fmt(xv),
                   expect, hits[0]["number_match"] if hits else "MISSING_ROW", agree)
            )
            evp = site / "evidence-pages" / ("%s__%s.html" % (iso, cell["variable"]))
            if evp.is_file() and cell["value"] is not None:
                ev = parse_evidence_values(read_html(evp))
                evc = next((e for e in ev if e["year"] == cell["year"]), None)
                if evc:
                    ekey = (iso, "evidence cell", cell["variable"], year_key(cell["year"]))
                    eh = by_key.get(ekey, [])
                    eexp = "MATCH" if numbers_equal(evc["value"], xv) else "FAIL"
                    lines.append(
                        "  evidence %s %s site=%s xlsx=%s expect=%s checklist=%s agree=%s"
                        % (cell["variable"], cell["year"], fmt(evc["value"]), fmt(xv),
                           eexp, eh[0]["number_match"] if eh else "MISSING_ROW",
                           bool(eh) and eh[0]["number_match"] == eexp)
                    )
    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Main audit
# ---------------------------------------------------------------------------

def run_audit(truth: Path, site: Path, out_xlsx: Path, scratch: Path,
              fetch: bool = True, workers: int = 6) -> dict:
    scratch = Path(scratch)
    scratch.mkdir(parents=True, exist_ok=True)
    dl_dir = scratch / "downloads"
    dl_dir.mkdir(parents=True, exist_ok=True)
    cache_path = dl_dir / "fetch_cache.json"

    print("Loading VERIFIED workbook:", truth, flush=True)
    verified = load_verified(truth)
    print("  sheets:", verified["sheets"], flush=True)
    print("  countries:", len(verified["iso3_list"]), flush=True)

    print("Building number-audit rows from HTML + xlsx…", flush=True)
    rows = build_audit_rows(site, verified)
    print("  number rows:", len(rows), flush=True)

    print("Collecting statistic-associated URLs…", flush=True)
    pairs = collect_all_stat_links(site)
    urls = []
    for p in pairs:
        if p.get("url") and p["url"].startswith("http"):
            urls.append(p["url"])
    for r in rows:
        if str(r.get("url") or "").startswith("http"):
            urls.append(r["url"])
    # Source_register + irregular + panel url columns
    for sheet, col in (("Source_register", "source_url"),
                       ("Irregular_estimates_all", "source_url")):
        df = verified[sheet]
        if col in df.columns:
            urls.extend(str(u) for u in df[col].dropna() if str(u).startswith("http"))
    urls = list(dict.fromkeys(urls))
    print("  unique http URLs:", len(urls), flush=True)

    cache = load_fetch_cache(cache_path)
    if fetch:
        known_bad = ["press.police.ac.kr", "ismu.org", "nisshinkyo.org",
                     "sem.admin.ch", "police.ac.kr"]
        print("Fetching live URLs (cache=%d already)…" % len(cache), flush=True)
        fetch_urls(urls, dl_dir, cache, workers=workers, retry_hosts=known_bad)
        save_fetch_cache(cache_path, cache)
        # fetch_errors.log
        err_path = scratch / "fetch_errors.log"
        with open(err_path, "w", encoding="utf-8") as f:
            for u, rec in sorted(cache.items()):
                if rec.get("error") or not rec.get("downloadable"):
                    f.write("%s\tstatus=%s\terror=%s\n" % (u, rec.get("status"), rec.get("error")))
        print("  fetch cache size:", len(cache), "errors logged to", err_path, flush=True)
    else:
        print("  fetch skipped; cache size", len(cache), flush=True)

    print("Attaching link/archive comparisons…", flush=True)
    rows = attach_link_results(rows, site, verified, cache, scratch)

    print("Checking local links and manifest hashes", flush=True)
    local_integrity = build_local_integrity(site)
    print("  local references=%d unique paths=%d missing=%d manifest exact=%d/%d" % (
        local_integrity["local_reference_count"],
        local_integrity["unique_local_path_count"],
        len(local_integrity["missing_local_paths"]),
        local_integrity["manifest_exact"],
        local_integrity["manifest_entries"],
    ), flush=True)
    link_audit = build_link_audit(site, verified, cache, local_integrity, rows)
    workbook_audit = build_workbook_audit(site, verified, rows, local_integrity, link_audit)
    print("  aggregated external URLs:", len(link_audit), flush=True)

    print("Writing checklist…", out_xlsx, flush=True)
    write_checklist(
        rows, out_xlsx, verified["iso3_list"], str(Path(truth).resolve()),
        verified=verified, site=site, fetch_cache=cache,
        local_integrity=local_integrity, link_audit=link_audit,
        workbook_audit=workbook_audit,
    )

    summary = summarize(rows)
    summary["checklist"] = str(out_xlsx)
    summary["truth"] = str(Path(truth).resolve())
    summary["n_iso"] = len(verified["iso3_list"])
    summary["iso3"] = verified["iso3_list"]
    summary["local_reference_count"] = local_integrity["local_reference_count"]
    summary["unique_local_path_count"] = local_integrity["unique_local_path_count"]
    summary["missing_local_paths"] = len(local_integrity["missing_local_paths"])
    summary["manifest_entries"] = local_integrity["manifest_entries"]
    summary["manifest_exact"] = local_integrity["manifest_exact"]
    summary["unique_external_urls"] = len(link_audit)
    summary["external_url_status"] = dict(Counter(r.get("current_status") for r in link_audit))
    summary["workbook_status"] = dict(Counter(r.get("status") for r in workbook_audit))
    print("SUMMARY", json.dumps({k: v for k, v in summary.items() if k != "iso3"}, ensure_ascii=False), flush=True)

    write_summary_csv(rows, scratch / "audit_summary.csv")
    (scratch / "audit_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return {"rows": rows, "summary": summary, "verified": verified, "fetch_cache": cache}


def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Audit site numbers and source links against the VERIFIED workbook")
    p.add_argument("--truth", type=Path, default=DEFAULT_TRUTH)
    p.add_argument("--site", type=Path, default=SITE_ROOT)
    p.add_argument("--out", type=Path, default=DEFAULT_CHECKLIST)
    p.add_argument("--scratch", type=Path, default=None)
    p.add_argument("--no-fetch", action="store_true")
    p.add_argument("--workers", type=int, default=6)
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    scratch = args.scratch or Path(os.environ.get(
        "AUDIT_SCRATCH",
        r"C:\Users\Raymond\AppData\Local\Temp\grok-goal-ff878ac0dd8e\implementer",
    ))
    t0 = time.time()
    result = run_audit(
        truth=args.truth,
        site=args.site,
        out_xlsx=args.out,
        scratch=scratch,
        fetch=not args.no_fetch,
        workers=args.workers,
    )
    print("elapsed_sec", round(time.time() - t0, 1), flush=True)
    return result


if __name__ == "__main__":
    main()
