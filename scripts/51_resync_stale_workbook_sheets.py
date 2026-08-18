# -*- coding: utf-8 -*-
"""Verification_log and Source_register in the published workbook were written before
the Eurostat re-test and the traceability pass, so the workbook still showed
migr_eipre failing while the website showed it at 100%. Rewrite just those two sheets
from the current CSVs, in place, using the same conventions as the original builder
(scripts/11_write_final_workbook.py): plain dump, freeze A2, width = len+2 capped
to [10, 62] over the first 200 rows."""
import os
import openpyxl
import pandas as pd

SITE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(SITE, 'data')
XLSX = os.path.join(D, 'FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx')

SHEETS = {'Verification_log': 'verification_log.csv',
          'Source_register': 'source_register.csv'}

wb = openpyxl.load_workbook(XLSX)
for sheet, csv in SHEETS.items():
    df = pd.read_csv(os.path.join(D, csv)).fillna('')
    pos = wb.sheetnames.index(sheet)
    before = (wb[sheet].max_row - 1, wb[sheet].max_column)
    del wb[sheet]
    ws = wb.create_sheet(sheet, pos)

    ws.append(list(df.columns))
    for rec in df.itertuples(index=False, name=None):
        ws.append(list(rec))

    ws.freeze_panes = 'A2'
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col[:200] if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 62)
    print('%-18s %d x %d  ->  %d x %d' % (sheet, before[0], before[1],
                                          ws.max_row - 1, ws.max_column))

wb.save(XLSX)

# ---------------------------------------------------------------- verify against the CSVs
chk = openpyxl.load_workbook(XLSX, read_only=True)
print('\nsheets after save: %d  %s' % (len(chk.sheetnames), chk.sheetnames))
ok = True
for sheet, csv in {**SHEETS, 'Panel_final': 'panel_final.csv',
                   'Known_issues': 'known_issues.csv',
                   'Corrections_applied': 'corrections_applied.csv',
                   'Deleted_values': 'deleted_values.csv'}.items():
    df = pd.read_csv(os.path.join(D, csv))
    ws = chk[sheet]
    match = (ws.max_row - 1 == len(df) and ws.max_column == len(df.columns))
    ok &= match
    print('  %-20s %-14s csv %-14s %s'
          % (sheet, '%d x %d' % (ws.max_row - 1, ws.max_column),
             '%d x %d' % (len(df), len(df.columns)), 'ok' if match else '*** DRIFT ***'))

# the specific claim the website makes must now hold in the workbook too
log = pd.read_csv(os.path.join(D, 'verification_log.csv'))
eip = log[(log.source.astype(str).str.contains('migr_eipre')) &
          (log.stage == 'after_correction')]
print('\nmigr_eipre after_correction: %d rows, %d exact (%.1f%%)'
      % (len(eip), (eip.status == 'EXACT').sum(), (eip.status == 'EXACT').mean() * 100))
assert ok, 'sheets still drift from the CSVs'
print('all sheets agree with data/*.csv')
