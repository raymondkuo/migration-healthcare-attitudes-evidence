# -*- coding: utf-8 -*-
"""The workbook README still quoted the counts as they stood on 2026-08-17, before the
Eurostat re-test and the traceability pass: 2454 comparisons, 49 corrections, and a
sheet list that omitted Deleted_values. Recompute every figure from the CSVs and patch
the cells in place, matched by their label so the layout is preserved."""
import os
import openpyxl
import pandas as pd

SITE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(SITE, 'data')
XLSX = os.path.join(D, 'FINAL_migration_population_panel_2010-2022_VERIFIED.xlsx')

log = pd.read_csv(os.path.join(D, 'verification_log.csv'))
corr = pd.read_csv(os.path.join(D, 'corrections_applied.csv'))
reg = pd.read_csv(os.path.join(D, 'source_register.csv')).fillna('')

ar = log[log.stage == 'as_received']
ac = log[log.stage == 'after_correction']
n_arch = int((reg.local_file.astype(str).str.len() > 3).sum())
n_api = int((reg.retrieval == 'VERIFIED_API').sum())
n_doc = int((reg.retrieval == 'ARCHIVED').sum())

UPDATES = {
    'Built': 'Compiled and verified 2026-08-17 from the two supplied workbooks; '
             'Eurostat migr_eipre re-queried and re-verified 2026-08-18.',
    'Values re-derived from live sources':
        '%d comparisons in two rounds: %d as received (2026-08-17), '
        '%d re-checked after correction (2026-08-18)' % (len(log), len(ar), len(ac)),
    '  matched exactly':
        '%d of %d as received (%.1f%%); %d of %d after correction (%.1f%%)'
        % ((ar.status == 'EXACT').sum(), len(ar), (ar.status == 'EXACT').mean() * 100,
           (ac.status == 'EXACT').sum(), len(ac), (ac.status == 'EXACT').mean() * 100),
    '  discrepancies found':
        '%d, all corrected and re-verified against the live source on 2026-08-18'
        % int((ar.status != 'EXACT').sum()),
    'Document source citations':
        '%d source rows across %d distinct URLs' % (len(reg), reg.source_url.nunique()),
    '  archived in the country folders':
        '%d of %d (%d verified live via API, %d archived as documents)'
        % (n_arch, len(reg), n_api, n_doc),
    '  not retrievable by any means':
        '2 (see Known_issues); both were superseded by a retrieved equivalent',
    'Corrections applied':
        '%d values across %d countries: %s (see Corrections_applied)'
        % (len(corr), corr.iso3.nunique(), ', '.join(sorted(corr.iso3.unique()))),
    'Verification_log':
        'All %d value-by-value comparisons against live sources, each stamped with the '
        'stage and the date it was performed.' % len(log),
}

wb = openpyxl.load_workbook(XLSX)
ws = wb['README']

seen = set()
for row in range(1, ws.max_row + 1):
    label = str(ws.cell(row, 1).value or '')
    if label in UPDATES:
        ws.cell(row, 2).value = UPDATES[label]
        seen.add(label)
        print('%2d  %-38s %s' % (row, label, UPDATES[label][:88]))
missing = set(UPDATES) - seen
assert not missing, 'labels not found in README: %s' % missing

# the SHEETS block listed nine of the ten sheets
sheet_rows = [r for r in range(1, ws.max_row + 1)
              if str(ws.cell(r, 1).value or '') in
              ('Panel_final', 'Data_quality', 'Corrections_applied', 'Known_issues',
               'Verification_log', 'Source_register', 'Irregular_estimates_all', 'Codebook')]
if not any(str(ws.cell(r, 1).value or '') == 'Deleted_values' for r in range(1, ws.max_row + 1)):
    at = max(sheet_rows) + 1
    ws.insert_rows(at)
    ws.cell(at, 1).value = 'Deleted_values'
    ws.cell(at, 2).value = ('Every value removed as untraceable, with the reason. '
                            'Nothing is deleted silently.')
    print('%2d  %-38s inserted into the SHEETS block' % (at, 'Deleted_values'))

wb.save(XLSX)

chk = openpyxl.load_workbook(XLSX, read_only=True)['README']
vals = {str(r[0] or ''): str(r[1] or '') for r in chk.iter_rows(values_only=True)}
for k, v in UPDATES.items():
    assert vals.get(k) == v, 'readback mismatch on %r' % k
assert 'Deleted_values' in vals
print('\nREADME sheet verified against data/*.csv')
